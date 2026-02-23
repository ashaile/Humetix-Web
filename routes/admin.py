import hmac
import logging
import os
import re
from datetime import datetime
from io import BytesIO

from flask import (
    Blueprint,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    send_from_directory,
    url_for,
)
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Border, Font, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage, ImageOps
from sqlalchemy import or_
from sqlalchemy.orm import joinedload

from models import Application, Inquiry, db
from routes.utils import BASE_DIR, ENV_FILE_PATH, UPLOAD_DIR, require_admin
from services.excel_service import (
    EXCEL_COLUMN_LABELS,
    _excel_row_values,
    parse_excel_columns as _parse_excel_columns,
)

logger = logging.getLogger(__name__)

admin_bp = Blueprint('admin', __name__)


def _update_env_value(env_path: str, key: str, value: str) -> None:
    lines = []
    found = False

    if os.path.exists(env_path):
        with open(env_path, "r", encoding="utf-8") as f:
            lines = f.readlines()

    for idx, line in enumerate(lines):
        stripped = line.strip()
        if stripped.startswith("#"):
            continue
        if line.startswith(f"{key}="):
            lines[idx] = f"{key}={value}\n"
            found = True
            break

    if not found:
        if lines and not lines[-1].endswith("\n"):
            lines[-1] += "\n"
        lines.append(f"{key}={value}\n")

    tmp_path = f"{env_path}.tmp"
    with open(tmp_path, "w", encoding="utf-8", newline="\n") as f:
        f.writelines(lines)
    os.replace(tmp_path, env_path)

def build_filtered_query(args):
    search_type = args.get('type', 'name')
    search_query = args.get('q', '')

    filters = {
        'gender': args.get('gender'),
        'shift': args.get('shift'),
        'posture': args.get('posture'),
        'overtime': args.get('overtime'),
        'holiday': args.get('holiday'),
        'agree': args.get('agree'),
        'advance_pay': args.get('advance_pay'),
        'insurance_type': args.get('insurance_type'),
        'status': args.get('status'),
    }

    start_date = args.get('start_date')
    end_date = args.get('end_date')

    query = Application.query.options(joinedload(Application.careers))

    if search_query:
        if search_type == 'name':
            query = query.filter(Application.name.contains(search_query))
        elif search_type == 'phone':
            query = query.filter(Application.phone.contains(search_query))

    if filters['gender']:
        query = query.filter(Application.gender == filters['gender'])
    if filters['shift']:
        query = query.filter(Application.shift == filters['shift'])
    if filters['posture']:
        query = query.filter(Application.posture == filters['posture'])
    if filters['overtime']:
        query = query.filter(Application.overtime == filters['overtime'])
    if filters['holiday']:
        query = query.filter(Application.holiday == filters['holiday'])
    if filters['agree']:
        is_agree = True if filters['agree'] == 'on' else False
        query = query.filter(Application.agree == is_agree)
    if filters['advance_pay']:
        query = query.filter(Application.advance_pay == filters['advance_pay'])
    if filters['insurance_type']:
        query = query.filter(Application.insurance_type == filters['insurance_type'])
    if filters['status']:
        query = query.filter(Application.status == filters['status'])

    if start_date:
        try:
            s_date = datetime.strptime(start_date, '%Y-%m-%d')
            query = query.filter(Application.timestamp >= s_date)
        except ValueError:
            start_date = ''
    if end_date:
        try:
            e_date = datetime.strptime(end_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
            query = query.filter(Application.timestamp <= e_date)
        except ValueError:
            end_date = ''

    return query, filters, search_query, start_date, end_date

@admin_bp.route('/humetix_master_99')
@require_admin
def master_view():
    query, filters, search_query, start_date, end_date = build_filtered_query(request.args)

    page = request.args.get('page', 1, type=int)
    per_page = 10
    pagination = query.order_by(Application.timestamp.desc()).paginate(page=page, per_page=per_page, error_out=False)

    data = [app.to_dict() for app in pagination.items]

    status_options = ['new', 'review', 'interview', 'offer', 'hired', 'rejected']
    return render_template(
        'admin.html',
        data=data,
        pagination=pagination,
        filters=filters,
        search_query=search_query,
        start_date=start_date,
        end_date=end_date,
        status_options=status_options,
        query_string=request.query_string.decode()
    )


@admin_bp.route('/admin/change-password', methods=['POST'])
@require_admin
def change_admin_password():
    current_password = request.form.get('current_password', '')
    new_password = request.form.get('new_password', '')
    confirm_password = request.form.get('confirm_password', '')
    existing_password = os.environ.get("ADMIN_PASSWORD", "")

    if not existing_password:
        logger.error("ADMIN_PASSWORD environment variable is not set")
        return jsonify({"success": False, "message": "서버 설정 오류가 발생했습니다."}), 500
    if not current_password or not hmac.compare_digest(current_password, existing_password):
        return jsonify({"success": False, "message": "현재 비밀번호가 올바르지 않습니다."}), 400
    if len(new_password) < 8:
        return jsonify({"success": False, "message": "새 비밀번호는 8자 이상이어야 합니다."}), 400
    if "\n" in new_password or "\r" in new_password:
        return jsonify({"success": False, "message": "새 비밀번호 형식이 올바르지 않습니다."}), 400
    if not hmac.compare_digest(new_password, confirm_password):
        return jsonify({"success": False, "message": "새 비밀번호 확인이 일치하지 않습니다."}), 400
    if hmac.compare_digest(new_password, existing_password):
        return jsonify({"success": False, "message": "기존과 다른 비밀번호를 입력해 주세요."}), 400

    try:
        _update_env_value(ENV_FILE_PATH, "ADMIN_PASSWORD", new_password)
        os.environ["ADMIN_PASSWORD"] = new_password
        logger.info("Admin password updated from admin page")
    except Exception:
        logger.exception("Failed to update ADMIN_PASSWORD")
        return jsonify({"success": False, "message": "비밀번호 저장 중 오류가 발생했습니다."}), 500

    return jsonify({"success": True, "message": "관리자 비밀번호가 변경되었습니다."})


@admin_bp.route('/download_excel')
@require_admin
def download_excel():
    query, _, _, _, _ = build_filtered_query(request.args)
    apps = query.order_by(Application.timestamp.desc()).all()

    selected_columns = _parse_excel_columns(request.args.get("excel_columns"))

    if selected_columns is not None:
        wb = Workbook()
        ws = wb.active
        ws.title = "지원자목록"

        headers = [EXCEL_COLUMN_LABELS[key] for key in selected_columns]
        ws.append(headers)
        for app in apps:
            values = _excel_row_values(app)
            ws.append([values.get(key, "") for key in selected_columns])

        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font

        ws.freeze_panes = "A2"
        sampled_row_count = min(ws.max_row, 200)
        for idx, key in enumerate(selected_columns, start=1):
            max_len = len(EXCEL_COLUMN_LABELS[key])
            for row_num in range(2, sampled_row_count + 1):
                cell_value = ws.cell(row=row_num, column=idx).value
                max_len = max(max_len, len(str(cell_value or "")))
            ws.column_dimensions[get_column_letter(idx)].width = min(max(12, max_len + 2), 40)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            as_attachment=True,
            download_name="humetix_applications.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    template_path = os.path.join(BASE_DIR, "templates", "excel", "입사지원서.xlsx")
    wb = load_workbook(template_path)
    template_ws = wb.active

    def safe_sheet_title(base):
        cleaned = "".join(ch for ch in base if ch not in r":\\/?*[]")
        cleaned = cleaned.strip()
        return cleaned[:31] if len(cleaned) > 31 else cleaned

    def set_value(ws, cell, value):
        ws[cell].value = value if value is not None else ""

    def parse_vision_value(vision_text):
        if not vision_text:
            return ""
        m = re.search(r"(\d+(?:[\.,]\d+)?)", str(vision_text))
        if not m:
            cleaned = re.sub(r"[^0-9\.,]", "", str(vision_text))
            if not cleaned:
                return ""
            m = re.search(r"(\d+(?:[\.,]\d+)?)", cleaned)
            if not m:
                return ""
        val = m.group(1)
        return val.replace(".", ",")

    def parse_vision_type(vision_text):
        if not vision_text:
            return ""
        text = str(vision_text)
        if "교정" in text:
            return "교정"
        if "나안" in text:
            return "나안"
        return ""

    for app in apps:
        date_str = app.timestamp.strftime("%Y-%m-%d") if app.timestamp else "0000-00-00"
        name = app.name or ""
        title = safe_sheet_title(f"[{date_str}]{name}")

        ws = wb.copy_worksheet(template_ws)
        ws.title = title
        if hasattr(ws, "_images"):
            ws._images = []
        # Slightly tighten column widths for readability
        for col_key, dim in ws.column_dimensions.items():
            if dim.width and dim.width > 1.5:
                dim.width = round(dim.width * 0.9, 2)

        # Basic info
        if name:
            # Split label/name across merged cells to avoid rich-text write errors
            for rng in list(ws.merged_cells.ranges):
                if rng.coord == "O4:AB4":
                    ws.unmerge_cells("O4:AB4")
                    break
            ws.merge_cells("O4:Q4")
            ws.merge_cells("R4:AB4")
            set_value(ws, "O4", "(한글)")
            ws["O4"].font = Font(size=14)
            set_value(ws, "R4", name)
            ws["R4"].font = Font(size=24, bold=True)
            # Name cell border: top double, left none, bottom/right thin
            top = Side(style="double")
            bottom = Side(style="thin")
            right = Side(style="thin")
            name_border = Border(top=top, bottom=bottom, right=right)
            for row in ws["R4:AB4"]:
                for cell in row:
                    cell.border = name_border
        else:
            set_value(ws, "O4", "")
        if app.birth:
            today = datetime.now().date()
            age = today.year - app.birth.year - ((today.month, today.day) < (app.birth.month, app.birth.day))
            birth_display = f"{app.birth.strftime('%Y-%m-%d')} ({age})"
        else:
            birth_display = ""
        set_value(ws, "AG4", birth_display)
        set_value(ws, "O6", app.phone or "")
        set_value(ws, "AG6", "")
        set_value(ws, "O7", app.email or "")
        set_value(ws, "O8", app.address or "")

        # Education (leave blank if not provided)
        set_value(ws, "O11", "")
        set_value(ws, "AK11", "")
        set_value(ws, "O12", "")
        set_value(ws, "AK12", "")

        # Careers (max 3 rows)
        career_rows = [15, 16, 17]
        for idx, row in enumerate(career_rows):
            if idx < len(app.careers):
                c = app.careers[idx]
                period = f"{c.start.strftime('%Y-%m-%d') if c.start else ''}~{c.end.strftime('%Y-%m-%d') if c.end else ''}"
                set_value(ws, f"D{row}", c.company or "")
                set_value(ws, f"O{row}", period)
                set_value(ws, f"AC{row}", c.role or "")
                set_value(ws, f"AK{row}", c.reason or "")
            else:
                set_value(ws, f"D{row}", "")
                set_value(ws, f"O{row}", "")
                set_value(ws, f"AC{row}", "")
                set_value(ws, f"AK{row}", "")
        if len(app.careers) > 3:
            set_value(ws, "D17", f"{app.careers[2].company or ''} 외 {len(app.careers) - 2}건")

        # Body / size
        set_value(ws, "K23", app.tshirt or "")
        set_value(ws, "O23", "")
        vision_val = parse_vision_value(app.vision)
        set_value(ws, "U23", vision_val)
        set_value(ws, "Z23", vision_val)
        vision_type = parse_vision_type(app.vision)
        if vision_type:
            set_value(ws, "R22", f"시 력 ( 나안 , 교정 )-{vision_type}")
        set_value(ws, "AC23", f"{app.height}cm" if app.height is not None else "")
        set_value(ws, "AG23", f"{app.weight}kg" if app.weight is not None else "")
        set_value(ws, "AK23", f"{app.shoes}mm" if app.shoes is not None else "")

        # Work conditions / confirmations
        set_value(ws, "L28", app.shift or "")
        set_value(ws, "L29", app.overtime or "")
        set_value(ws, "L30", app.holiday or "")
        set_value(ws, "L31", app.posture or "")
        set_value(ws, "L32", "")
        set_value(ws, "AG32", app.start_date.strftime("%Y-%m-%d") if app.start_date else "")

        if app.photo:
            photo_path = os.path.join(UPLOAD_DIR, app.photo)
            if os.path.exists(photo_path):
                try:
                    try:
                        from pillow_heif import register_heif_opener
                        register_heif_opener()
                    except Exception:
                        pass
                    with PILImage.open(photo_path) as img:
                        img = ImageOps.exif_transpose(img)
                        if img.mode not in ('RGB',):
                            img = img.convert('RGB')
                        # estimate target size from merged range B4:H8
                        def col_width(col):
                            w = ws.column_dimensions[col].width
                            return w if w else 8.43
                        def row_height(r):
                            h = ws.row_dimensions[r].height
                            return h if h else 15
                        width_px = int(sum(col_width(c) * 7 for c in ["B","C","D","E","F","G","H"]))
                        height_px = int(sum(row_height(r) * 1.33 for r in range(4, 9)))
                        # keep higher source resolution to avoid blur
                        max_dim = (800, 1000)
                        if img.width > max_dim[0] or img.height > max_dim[1]:
                            img.thumbnail(max_dim)
                        from io import BytesIO as _BytesIO
                        img_stream = _BytesIO()
                        img.save(img_stream, format='PNG', optimize=False)
                        img_stream.seek(0)
                        excel_img = ExcelImage(img_stream)
                        excel_img.width = width_px
                        excel_img.height = height_px
                        ws.add_image(excel_img, "B4")
                except Exception:
                    pass

    if apps:
        wb.remove(template_ws)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="humetix_applications.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@admin_bp.route('/update_memo/<app_id>', methods=['POST'])
@require_admin
def update_memo(app_id):
    app = db.session.get(Application, app_id)
    if not app:
        return jsonify({"success": False, "message": "지원서를 찾을 수 없습니다."}), 404

    app.memo = request.form.get('memo', '')
    try:
        db.session.commit()
        return jsonify({"success": True})
    except Exception:
        db.session.rollback()
        return jsonify({"success": False, "message": "서버 오류가 발생했습니다."}), 500


@admin_bp.route('/update_status/<app_id>', methods=['POST'])
@require_admin
def update_status(app_id):
    app = db.session.get(Application, app_id)
    if not app:
        return jsonify({"success": False, "message": "지원서를 찾을 수 없습니다."}), 404

    status_val = request.form.get('status', '')
    allowed = {'new', 'review', 'interview', 'offer', 'hired', 'rejected'}
    if status_val not in allowed:
        return jsonify({"success": False, "message": "유효하지 않은 상태입니다."}), 400

    app.status = status_val
    try:
        db.session.commit()
        return jsonify({"success": True})
    except Exception:
        db.session.rollback()
        return jsonify({"success": False, "message": "상태 저장 오류가 발생했습니다."}), 500

@admin_bp.route('/inquiries')
@require_admin
def inquiries():
    q = request.args.get('q', '').strip()
    status = request.args.get('status', '').strip()

    query = Inquiry.query
    if q:
        like = f"%{q}%"
        query = query.filter(or_(
            Inquiry.company.ilike(like),
            Inquiry.name.ilike(like),
            Inquiry.phone.ilike(like),
            Inquiry.email.ilike(like),
        ))
    if status:
        query = query.filter(Inquiry.status == status)

    items = query.order_by(Inquiry.created_at.desc()).all()
    status_options = ['new', 'in_progress', 'done']
    return render_template(
        'admin_inquiries.html',
        items=items,
        q=q,
        status=status,
        status_options=status_options
    )

@admin_bp.route('/inquiries/delete', methods=['POST'])
@require_admin
def delete_inquiries():
    selected_ids = request.form.getlist('selected_ids')
    if not selected_ids:
        return redirect(url_for('admin.inquiries'))

    try:
        for inquiry_id in selected_ids:
            item = db.session.get(Inquiry, inquiry_id)
            if item:
                db.session.delete(item)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error deleting inquiries: {e}")

    return redirect(url_for('admin.inquiries'))


@admin_bp.route('/inquiries/update/<int:inquiry_id>', methods=['POST'])
@require_admin
def update_inquiry(inquiry_id):
    item = db.session.get(Inquiry, inquiry_id)
    if not item:
        return redirect(url_for('admin.inquiries'))

    item.status = request.form.get('status', item.status)
    item.assignee = request.form.get('assignee', item.assignee)
    item.admin_memo = request.form.get('admin_memo', item.admin_memo)
    db.session.commit()

    return redirect(url_for('admin.inquiries'))


@admin_bp.route('/delete_selected', methods=['POST'])
@require_admin
def delete_selected():
    selected_ids = request.form.getlist('selected_ids')
    if not selected_ids:
        return redirect(url_for('admin.master_view'))

    apps = Application.query.filter(Application.id.in_(selected_ids)).all()
    for app in apps:
        if app.photo:
            try:
                os.remove(os.path.join(UPLOAD_DIR, app.photo))
            except Exception as e:
                logger.error(f"Error deleting photo {app.photo}: {e}")
        db.session.delete(app)
    db.session.commit()

    return redirect(url_for('admin.master_view'))
@admin_bp.route('/view_photo/<filename>')
@require_admin
def view_photo(filename):
    # 경로 탈출 방지: 파일명에 디렉토리 구분자가 포함되면 거부
    if filename != os.path.basename(filename):
        return "Not Found", 404
    file_path = os.path.join(UPLOAD_DIR, filename)
    if not os.path.exists(file_path):
        return "Not Found", 404

    ext = os.path.splitext(filename)[1].lower()
    if ext in {'.heic', '.heif'}:
        try:
            from pillow_heif import register_heif_opener
            register_heif_opener()
            with PILImage.open(file_path) as img:
                if img.mode not in ('RGB',):
                    img = img.convert('RGB')
                buf = BytesIO()
                img.save(buf, format='JPEG', quality=85)
                buf.seek(0)
            return send_file(buf, mimetype='image/jpeg', download_name=f"{os.path.splitext(filename)[0]}.jpg")
        except Exception as e:
            logger.error(f"HEIC preview failed for {filename}: {e}")
            return "Unsupported image", 415

    return send_from_directory(UPLOAD_DIR, filename)

@admin_bp.route('/clear_data', methods=['POST'])
@require_admin
def clear_data():
    # 모든 지원서 삭제 (ORM 캐스케이드 적용) + 파일 정리
    apps = Application.query.all()
    referenced_photos = set()
    for app in apps:
        if app.photo:
            referenced_photos.add(app.photo)
            try:
                os.remove(os.path.join(UPLOAD_DIR, app.photo))
            except Exception as e:
                logger.error(f"Error deleting photo {app.photo}: {e}")
        db.session.delete(app)
    db.session.commit()
    
    # 남아있는 파일 정리 (DB에 없는 파일)
    for f in os.listdir(UPLOAD_DIR):
        if f not in referenced_photos:
            try:
                os.remove(os.path.join(UPLOAD_DIR, f))
            except Exception as e:
                logger.error(f"Error deleting leftover file {f}: {e}")
        
    return redirect(url_for('admin.master_view'))

