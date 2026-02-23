import logging
import re
from datetime import datetime
from io import BytesIO

from flask import (
    Blueprint,
    current_app,
    jsonify,
    render_template,
    request,
    send_file,
)
from models import Employee, Payslip, db
from routes.utils import require_admin, validate_month as _validate_month
from services.payslip_service import ALLOWED_SALARY_MODES, compute_payslips
from extensions import limiter

logger = logging.getLogger(__name__)

payslip_bp = Blueprint("payslip", __name__)


@payslip_bp.route("/admin/payslip")
@require_admin
def admin_payslip():
    month = request.args.get("month", datetime.now().strftime("%Y-%m"))
    if not _validate_month(month):
        month = datetime.now().strftime("%Y-%m")

    payslips = Payslip.query.filter(Payslip.month == month).order_by(Payslip.emp_name.asc()).all()

    cfg = current_app.config
    return render_template(
        "admin_payslip.html",
        payslips=payslips,
        month=month,
        salary_mode=cfg.get("SALARY_MODE", "standard"),
        hourly_wage=cfg.get("HOURLY_WAGE", 10320),
    )


@payslip_bp.route("/admin/payslip/generate", methods=["POST"])
@require_admin
def generate_payslips():
    payload = request.get_json(silent=True) or {}
    month = (request.form.get("month") or payload.get("month") or "").strip()
    if not _validate_month(month):
        return jsonify({"error": "month (YYYY-MM) is required"}), 400

    salary_mode = (
        request.form.get("salary_mode")
        or payload.get("salary_mode")
        or request.args.get("salary_mode")
        or current_app.config.get("SALARY_MODE", "standard")
    )
    if salary_mode not in ALLOWED_SALARY_MODES:
        return jsonify({"error": f"invalid salary_mode: {salary_mode}"}), 400

    try:
        result = compute_payslips(month, salary_mode)
    except Exception as exc:
        db.session.rollback()
        logger.error("Payslip generate error: %s", exc)
        return jsonify({"error": "급여 생성 중 오류가 발생했습니다."}), 500

    if isinstance(result, str):
        return jsonify({"error": result}), 404

    created, updated, skipped = result
    return jsonify({"success": True, "created": created, "updated": updated, "skipped": skipped})


@payslip_bp.route("/admin/payslip/<int:payslip_id>", methods=["DELETE"])
@require_admin
def delete_payslip(payslip_id: int):
    payslip = db.session.get(Payslip, payslip_id)
    if not payslip:
        return jsonify({"error": "급여명세서를 찾을 수 없습니다."}), 404

    try:
        db.session.delete(payslip)
        db.session.commit()
        return jsonify({"success": True, "deleted": 1, "id": payslip_id})
    except Exception as exc:
        db.session.rollback()
        logger.error("Payslip delete error: %s", exc)
        return jsonify({"error": "급여명세서 삭제 중 오류가 발생했습니다."}), 500


@payslip_bp.route("/admin/payslip/<int:payslip_id>", methods=["PUT"])
@require_admin
def update_payslip(payslip_id: int):
    """관리자 수동 급여 수정. is_manual=True 자동 설정."""
    payslip = db.session.get(Payslip, payslip_id)
    if not payslip:
        return jsonify({"error": "급여명세서를 찾을 수 없습니다."}), 404

    data = request.get_json(silent=True) or {}
    editable = [
        "base_salary", "ot_pay", "night_pay", "holiday_pay",
        "tax", "pension", "health_ins", "longterm_care",
        "employment_ins", "advance_deduction",
    ]
    try:
        for field in editable:
            if field in data:
                setattr(payslip, field, int(data[field]))

        # gross/insurance/net 자동 재계산
        payslip.gross = payslip.base_salary + payslip.ot_pay + payslip.night_pay + payslip.holiday_pay
        payslip.insurance = payslip.pension + payslip.health_ins + payslip.longterm_care + payslip.employment_ins
        payslip.net = payslip.gross - payslip.tax - payslip.insurance - payslip.advance_deduction
        payslip.is_manual = True

        db.session.commit()
        return jsonify({"success": True, "payslip": payslip.to_dict()})
    except Exception as exc:
        db.session.rollback()
        logger.error("Payslip update error: %s", exc)
        return jsonify({"error": "급여명세서 수정 중 오류가 발생했습니다."}), 500


@payslip_bp.route("/admin/payslip/<int:payslip_id>/reset", methods=["POST"])
@require_admin
def reset_payslip(payslip_id: int):
    """수동 수정 초기화 — 근태 데이터 기반 재계산."""
    payslip = db.session.get(Payslip, payslip_id)
    if not payslip:
        return jsonify({"error": "급여명세서를 찾을 수 없습니다."}), 404

    if not payslip.is_manual:
        return jsonify({"error": "수동 수정된 명세서가 아닙니다."}), 400

    from services.payslip_service import _month_range
    from models import AttendanceRecord, AdvanceRequest
    from sqlalchemy import func

    cfg = current_app.config
    hourly = cfg.get("HOURLY_WAGE", 10320)
    std_monthly_hours = cfg.get("MONTHLY_STANDARD_HOURS", 209)
    ot_mult = cfg.get("OT_MULTIPLIER", 1.5)
    night_prem = cfg.get("NIGHT_PREMIUM", 0.5)
    tax_rate = cfg.get("TAX_RATE", 0.033)
    pension_rate = cfg.get("PENSION_RATE", 0.045)
    health_rate = cfg.get("HEALTH_RATE", 0.03545)
    longterm_rate = cfg.get("LONGTERM_CARE_RATE", 0.1295)
    employment_rate = cfg.get("EMPLOYMENT_RATE", 0.009)

    start_date, end_date = _month_range(payslip.month)

    row = (
        db.session.query(
            func.sum(AttendanceRecord.total_work_hours).label("total_hours"),
            func.sum(AttendanceRecord.overtime_hours).label("ot_hours"),
            func.sum(AttendanceRecord.night_hours).label("night_hours"),
            func.sum(AttendanceRecord.holiday_work_hours).label("holiday_hours"),
        )
        .filter(
            AttendanceRecord.employee_id == payslip.employee_id,
            AttendanceRecord.work_date >= start_date,
            AttendanceRecord.work_date < end_date,
        )
        .first()
    )

    if not row or not row.total_hours:
        return jsonify({"error": "해당 월 근태 기록이 없어 초기화할 수 없습니다."}), 404

    total_h = round(row.total_hours or 0, 2)
    ot_h = round(row.ot_hours or 0, 2)
    night_h = round(row.night_hours or 0, 2)
    holiday_h = round(row.holiday_hours or 0, 2)

    if payslip.salary_mode == "actual":
        base_salary = round(hourly * total_h)
    else:
        base_salary = hourly * std_monthly_hours

    ot_pay = round(ot_h * hourly * ot_mult)
    night_pay = round(night_h * hourly * night_prem)
    holiday_pay = round(holiday_h * hourly * ot_mult)
    gross = base_salary + ot_pay + night_pay + holiday_pay

    # 직원별 보험 유형 조회
    emp = db.session.get(Employee, payslip.employee_id)
    emp_ins_type = emp.insurance_type if emp else "3.3%"

    if emp_ins_type == "4대보험":
        tax = 0
        pension = round(gross * pension_rate)
        health = round(gross * health_rate)
        longterm = round(health * longterm_rate)
        employment = round(gross * employment_rate)
    else:
        tax = round(gross * tax_rate)
        pension = 0
        health = 0
        longterm = 0
        employment = 0
    insurance = pension + health + longterm + employment

    adv_total = (
        db.session.query(func.coalesce(func.sum(AdvanceRequest.amount), 0))
        .filter(
            AdvanceRequest.employee_id == payslip.employee_id,
            AdvanceRequest.request_month == payslip.month,
            AdvanceRequest.status == "approved",
        )
        .scalar()
    )

    net = gross - tax - insurance - adv_total

    try:
        payslip.total_work_hours = total_h
        payslip.ot_hours = ot_h
        payslip.night_hours = night_h
        payslip.holiday_hours = holiday_h
        payslip.base_salary = base_salary
        payslip.ot_pay = ot_pay
        payslip.night_pay = night_pay
        payslip.holiday_pay = holiday_pay
        payslip.gross = gross
        payslip.tax = tax
        payslip.pension = pension
        payslip.health_ins = health
        payslip.longterm_care = longterm
        payslip.employment_ins = employment
        payslip.insurance = insurance
        payslip.advance_deduction = adv_total
        payslip.net = net
        payslip.is_manual = False
        db.session.commit()
        return jsonify({"success": True, "payslip": payslip.to_dict()})
    except Exception as exc:
        db.session.rollback()
        logger.error("Payslip reset error: %s", exc)
        return jsonify({"error": "초기화 중 오류가 발생했습니다."}), 500


@payslip_bp.route("/admin/payslip/delete-month", methods=["POST"])
@require_admin
def delete_payslips_by_month():
    payload = request.get_json(silent=True) or {}
    month = (request.form.get("month") or payload.get("month") or "").strip()
    if not _validate_month(month):
        return jsonify({"error": "month (YYYY-MM) is required"}), 400

    try:
        deleted_count = (
            Payslip.query.filter(Payslip.month == month).delete(synchronize_session=False)
        )
        db.session.commit()
        return jsonify({"success": True, "deleted": deleted_count, "month": month})
    except Exception as exc:
        db.session.rollback()
        logger.error("Payslip month delete error: %s", exc)
        return jsonify({"error": "월 급여명세서 삭제 중 오류가 발생했습니다."}), 500


def _generate_payslip_pdf(payslips):
    """급여명세서 PDF 생성 (공통 헬퍼). BytesIO 반환."""
    import os

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import SimpleDocTemplate, Spacer, Table, TableStyle, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

    buf = BytesIO()

    font_name = "Helvetica"
    font_name_bold = "Helvetica-Bold"
    for font_path in [
        "C:/Windows/Fonts/malgun.ttf",
        "C:/Windows/Fonts/malgunbd.ttf",
        "/usr/share/fonts/truetype/nanum/NanumGothic.ttf",
        "/usr/share/fonts/truetype/nanum/NanumGothicBold.ttf",
    ]:
        if os.path.exists(font_path):
            try:
                fname = os.path.basename(font_path).replace(".ttf", "").replace(".ttc", "")
                pdfmetrics.registerFont(TTFont(fname, font_path))
                if "bold" in fname.lower() or "bd" in fname.lower():
                    font_name_bold = fname
                else:
                    font_name = fname
            except Exception:
                continue

    hourly = current_app.config.get("HOURLY_WAGE", 10320)

    # 색상 정의
    HEADER_BG = colors.HexColor("#E85D26")
    HEADER_TEXT = colors.white
    SECTION_BG = colors.HexColor("#F8F9FA")
    TOTAL_BG = colors.HexColor("#FFF3ED")
    NET_BG = colors.HexColor("#E85D26")
    BORDER_COLOR = colors.HexColor("#DEE2E6")
    TEXT_COLOR = colors.HexColor("#212529")
    DEDUCT_COLOR = colors.HexColor("#DC2626")

    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=25 * mm, rightMargin=25 * mm,
        topMargin=20 * mm, bottomMargin=20 * mm,
    )

    elements = []

    for idx, payslip in enumerate(payslips):
        if idx > 0:
            elements.append(Spacer(1, 0))  # pagebreak handled by KeepTogether

        # ── 회사명 헤더 ──
        header_data = [["Humetix Inc."]]
        header_table = Table(header_data, colWidths=[160 * mm])
        header_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), font_name_bold),
            ("FONTSIZE", (0, 0), (-1, -1), 18),
            ("TEXTCOLOR", (0, 0), (-1, -1), HEADER_BG),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
        ]))
        elements.append(header_table)

        # 구분선
        line_data = [["" ]]
        line_table = Table(line_data, colWidths=[160 * mm], rowHeights=[2])
        line_table.setStyle(TableStyle([
            ("LINEBELOW", (0, 0), (-1, -1), 2, HEADER_BG),
        ]))
        elements.append(line_table)
        elements.append(Spacer(1, 4 * mm))

        # ── 제목 ──
        year, mon = payslip.month.split("-")
        title_data = [[f"{year}년 {int(mon)}월 급여명세서"]]
        title_table = Table(title_data, colWidths=[160 * mm])
        title_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), font_name_bold),
            ("FONTSIZE", (0, 0), (-1, -1), 16),
            ("TEXTCOLOR", (0, 0), (-1, -1), TEXT_COLOR),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(title_table)
        elements.append(Spacer(1, 3 * mm))

        # ── 직원 정보 박스 ──
        mode_label = "209h 고정" if payslip.salary_mode == "standard" else "실근무시간"
        info_data = [
            ["성명", payslip.emp_name, "부서", payslip.dept or "-"],
            ["시급", f"{hourly:,}원", "계산방식", mode_label],
            ["총근무시간", f"{payslip.total_work_hours}h", "잔업", f"{payslip.ot_hours}h"],
            ["심야", f"{payslip.night_hours}h", "휴일", f"{payslip.holiday_hours}h"],
        ]
        info_table = Table(info_data, colWidths=[30 * mm, 50 * mm, 30 * mm, 50 * mm])
        info_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("FONTNAME", (0, 0), (0, -1), font_name_bold),
            ("FONTNAME", (2, 0), (2, -1), font_name_bold),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#6B7280")),
            ("TEXTCOLOR", (2, 0), (2, -1), colors.HexColor("#6B7280")),
            ("TEXTCOLOR", (1, 0), (1, -1), TEXT_COLOR),
            ("TEXTCOLOR", (3, 0), (3, -1), TEXT_COLOR),
            ("BACKGROUND", (0, 0), (-1, -1), SECTION_BG),
            ("BOX", (0, 0), (-1, -1), 1, BORDER_COLOR),
            ("INNERGRID", (0, 0), (-1, -1), 0.5, BORDER_COLOR),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 6 * mm))

        # ── 지급 내역 ──
        earn_header = [["지급 내역", ""]]
        earn_data = [
            ["기본급", f"{payslip.base_salary:,}원"],
            [f"잔업수당 ({payslip.ot_hours}h)", f"{payslip.ot_pay:,}원"],
            [f"심야수당 ({payslip.night_hours}h)", f"{payslip.night_pay:,}원"],
            [f"휴일수당 ({payslip.holiday_hours}h)", f"{payslip.holiday_pay:,}원"],
        ]
        earn_total = [["총지급액", f"{payslip.gross:,}원"]]

        earn_all = earn_header + earn_data + earn_total
        earn_table = Table(earn_all, colWidths=[100 * mm, 60 * mm])
        earn_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("TEXTCOLOR", (0, 0), (-1, -1), TEXT_COLOR),
            # 헤더 행
            ("FONTNAME", (0, 0), (-1, 0), font_name_bold),
            ("FONTSIZE", (0, 0), (-1, 0), 12),
            ("BACKGROUND", (0, 0), (-1, 0), HEADER_BG),
            ("TEXTCOLOR", (0, 0), (-1, 0), HEADER_TEXT),
            # 합계 행
            ("FONTNAME", (0, -1), (-1, -1), font_name_bold),
            ("BACKGROUND", (0, -1), (-1, -1), TOTAL_BG),
            ("FONTSIZE", (0, -1), (-1, -1), 11),
            # 금액 오른쪽 정렬
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            # 테두리
            ("BOX", (0, 0), (-1, -1), 1, BORDER_COLOR),
            ("LINEBELOW", (0, 0), (-1, -2), 0.5, BORDER_COLOR),
            ("LINEABOVE", (0, -1), (-1, -1), 1, BORDER_COLOR),
            ("TOPPADDING", (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ("LEFTPADDING", (0, 0), (-1, -1), 10),
            ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ]))
        elements.append(earn_table)
        elements.append(Spacer(1, 5 * mm))

        # ── 공제 내역 ──
        ded_header = [["공제 내역", ""]]
        ded_data = []
        if payslip.tax > 0:
            ded_data.append(["소득세 (3.3%)", f"-{payslip.tax:,}원"])
        if payslip.pension > 0:
            ded_data.append(["국민연금 (4.75%)", f"-{payslip.pension:,}원"])
        if payslip.health_ins > 0:
            ded_data.append(["건강보험 (3.595%)", f"-{payslip.health_ins:,}원"])
        if payslip.longterm_care > 0:
            ded_data.append(["장기요양보험", f"-{payslip.longterm_care:,}원"])
        if payslip.employment_ins > 0:
            ded_data.append(["고용보험 (1.15%)", f"-{payslip.employment_ins:,}원"])
        if payslip.advance_deduction > 0:
            ded_data.append(["가불 차감", f"-{payslip.advance_deduction:,}원"])

        total_deduct = payslip.tax + payslip.insurance + payslip.advance_deduction
        ded_total = [["공제합계", f"-{total_deduct:,}원"]]

        ded_all = ded_header + ded_data + ded_total
        ded_table = Table(ded_all, colWidths=[100 * mm, 60 * mm])
        ded_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("TEXTCOLOR", (0, 1), (-1, -1), TEXT_COLOR),
            ("TEXTCOLOR", (1, 1), (1, -1), DEDUCT_COLOR),
            # 헤더 행
            ("FONTNAME", (0, 0), (-1, 0), font_name_bold),
            ("FONTSIZE", (0, 0), (-1, 0), 12),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#6B7280")),
            ("TEXTCOLOR", (0, 0), (-1, 0), HEADER_TEXT),
            # 합계 행
            ("FONTNAME", (0, -1), (-1, -1), font_name_bold),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#FEF2F2")),
            ("FONTSIZE", (0, -1), (-1, -1), 11),
            # 금액 오른쪽 정렬
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            # 테두리
            ("BOX", (0, 0), (-1, -1), 1, BORDER_COLOR),
            ("LINEBELOW", (0, 0), (-1, -2), 0.5, BORDER_COLOR),
            ("LINEABOVE", (0, -1), (-1, -1), 1, BORDER_COLOR),
            ("TOPPADDING", (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ("LEFTPADDING", (0, 0), (-1, -1), 10),
            ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ]))
        elements.append(ded_table)
        elements.append(Spacer(1, 6 * mm))

        # ── 실수령액 ──
        net_data = [[f"실수령액        {payslip.net:,}원"]]
        net_table = Table(net_data, colWidths=[160 * mm])
        net_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), font_name_bold),
            ("FONTSIZE", (0, 0), (-1, -1), 16),
            ("BACKGROUND", (0, 0), (-1, -1), NET_BG),
            ("TEXTCOLOR", (0, 0), (-1, -1), HEADER_TEXT),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("TOPPADDING", (0, 0), (-1, -1), 12),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
            ("BOX", (0, 0), (-1, -1), 1, HEADER_BG),
        ]))
        elements.append(net_table)
        elements.append(Spacer(1, 8 * mm))

        # ── 발급일 ──
        date_data = [[f"발급일: {datetime.now().strftime('%Y-%m-%d')}"]]
        date_table = Table(date_data, colWidths=[160 * mm])
        date_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#9CA3AF")),
            ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
        ]))
        elements.append(date_table)

        # 수동 수정 표시
        if payslip.is_manual:
            manual_data = [["* 본 명세서는 관리자에 의해 수동 수정되었습니다."]]
            manual_table = Table(manual_data, colWidths=[160 * mm])
            manual_table.setStyle(TableStyle([
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#F59E0B")),
                ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
            ]))
            elements.append(manual_table)

        # 페이지 나누기 (마지막 제외)
        if idx < len(payslips) - 1:
            from reportlab.platypus import PageBreak
            elements.append(PageBreak())

    doc.build(elements)
    buf.seek(0)
    return buf


@payslip_bp.route("/admin/payslip/pdf")
@require_admin
def payslip_pdf():
    month = request.args.get("month", datetime.now().strftime("%Y-%m"))
    if not _validate_month(month):
        return jsonify({"error": "invalid month format"}), 400

    employee_id = request.args.get("employee_id") or request.args.get("emp_id")
    query = Payslip.query.filter(Payslip.month == month)
    if employee_id:
        if not str(employee_id).isdigit():
            return jsonify({"error": "employee_id must be integer"}), 400
        query = query.filter(Payslip.employee_id == int(employee_id))

    payslips = query.order_by(Payslip.emp_name.asc()).all()
    if not payslips:
        return jsonify({"error": "해당 월 급여 데이터가 없습니다."}), 404

    buf = _generate_payslip_pdf(payslips)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"급여명세서_{month}.pdf",
        mimetype="application/pdf",
    )


@payslip_bp.route("/admin/payslip/excel")
@require_admin
def payslip_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font

    month = request.args.get("month", datetime.now().strftime("%Y-%m"))
    if not _validate_month(month):
        return jsonify({"error": "invalid month format"}), 400

    payslips = Payslip.query.filter(Payslip.month == month).order_by(Payslip.emp_name.asc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = f"{month} 급여"

    headers = [
        "직원ID",
        "이름",
        "부서",
        "계산방식",
        "총근무(h)",
        "잔업(h)",
        "야간(h)",
        "휴일(h)",
        "기본급",
        "잔업수당",
        "야간수당",
        "휴일수당",
        "총지급",
        "소득세",
        "국민연금",
        "건강보험",
        "장기요양",
        "고용보험",
        "4대보험합계",
        "가불차감",
        "실수령액",
        "수동수정",
    ]
    bold = Font(bold=True)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = bold

    for i, payslip in enumerate(payslips, 2):
        mode_label = "209h고정" if payslip.salary_mode == "standard" else "실근무"
        values = [
            payslip.employee_id,
            payslip.emp_name,
            payslip.dept,
            mode_label,
            payslip.total_work_hours,
            payslip.ot_hours,
            payslip.night_hours,
            payslip.holiday_hours,
            payslip.base_salary,
            payslip.ot_pay,
            payslip.night_pay,
            payslip.holiday_pay,
            payslip.gross,
            payslip.tax,
            payslip.pension,
            payslip.health_ins,
            payslip.longterm_care,
            payslip.employment_ins,
            payslip.insurance,
            payslip.advance_deduction,
            payslip.net,
            "Y" if payslip.is_manual else "",
        ]
        for col, value in enumerate(values, 1):
            ws.cell(row=i, column=col, value=value)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"급여명세서_{month}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ── Public payslip lookup ──


@payslip_bp.route("/payslip", methods=["GET", "POST"])
@limiter.limit("10 per minute")
def payslip_lookup():
    if request.method == "GET":
        return render_template("payslip_lookup.html")

    birth_date = request.form.get("birth_date", "").strip()
    emp_name = request.form.get("emp_name", "").strip()

    errors = []
    if not re.fullmatch(r"\d{6}", birth_date):
        errors.append("생년월일은 YYMMDD 6자리 숫자여야 합니다.")
    if not emp_name:
        errors.append("이름을 입력해주세요.")
    if errors:
        return render_template("payslip_lookup.html", errors=errors, form=request.form)

    employee = Employee.query.filter_by(
        name=emp_name, birth_date=birth_date, is_active=True
    ).first()
    if not employee:
        return render_template(
            "payslip_lookup.html",
            errors=["등록된 재직 직원이 아닙니다."],
            form=request.form,
        )

    payslips = (
        Payslip.query.filter_by(employee_id=employee.id)
        .order_by(Payslip.month.desc())
        .all()
    )
    return render_template(
        "payslip_lookup.html",
        payslips=payslips,
        employee=employee,
        form=request.form,
    )


@payslip_bp.route("/payslip/pdf")
@limiter.limit("10 per minute")
def payslip_public_pdf():
    birth_date = request.args.get("birth_date", "").strip()
    emp_name = request.args.get("emp_name", "").strip()
    month = request.args.get("month", "").strip()

    if not (re.fullmatch(r"\d{6}", birth_date) and emp_name and _validate_month(month)):
        return jsonify({"error": "잘못된 요청입니다."}), 400

    employee = Employee.query.filter_by(
        name=emp_name, birth_date=birth_date, is_active=True
    ).first()
    if not employee:
        return jsonify({"error": "직원 정보를 확인할 수 없습니다."}), 403

    payslip = Payslip.query.filter_by(
        employee_id=employee.id, month=month
    ).first()
    if not payslip:
        return jsonify({"error": "해당 월 급여 데이터가 없습니다."}), 404

    buf = _generate_payslip_pdf([payslip])
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"급여명세서_{month}_{emp_name}.pdf",
        mimetype="application/pdf",
    )
