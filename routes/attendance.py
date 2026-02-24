import logging
import os
import re
from calendar import monthrange
from datetime import date, datetime
from io import BytesIO

from flask import (
    Blueprint,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)
from sqlalchemy.exc import IntegrityError, OperationalError

from extensions import limiter
from models import AttendanceRecord, Employee, OperationCalendarDay, Site, db
from routes.utils import require_admin
from services.attendance_service import (
    ALLOWED_WORK_TYPES,
    CALENDAR_DAY_TYPES,
    LEAVE_DAY_TYPES,
    TIME_REQUIRED_TYPES,
    _default_day_type,
    _effective_day_type,
    _get_cfg,
    _parse_date,
    _validate_hhmm,
    calc_work_hours,
)

logger = logging.getLogger(__name__)

attendance_bp = Blueprint("attendance", __name__)


def _validate_month(value: str) -> bool:
    if not value or not re.fullmatch(r"\d{4}-\d{2}", value):
        return False
    year, month = map(int, value.split("-"))
    return 2000 <= year <= 2100 and 1 <= month <= 12


def _month_bounds(month_text: str):
    year, month = map(int, month_text.split("-"))
    last_day = monthrange(year, month)[1]
    return date(year, month, 1), date(year, month, last_day)


def _db_not_ready_message():
    return (
        "데이터베이스 초기화가 필요합니다. "
        "프로젝트 폴더에서 python fix_db.py 실행 후 서버를 재시작하세요."
    )


def _db_not_ready_json():
    return jsonify({"error": _db_not_ready_message()}), 503


def _db_not_ready_page():
    return render_template(
        "error.html",
        error_code="503",
        error_message="데이터베이스 초기화 필요",
        error_description=_db_not_ready_message(),
    ), 503


@attendance_bp.route("/api/attendance", methods=["POST"])
@limiter.limit("10 per minute")
def create_attendance():
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "JSON body required"}), 400

    required = ["emp_name", "birth_date", "work_date"]
    missing = [field for field in required if not data.get(field)]
    if missing:
        return jsonify({"error": f"Missing fields: {', '.join(missing)}"}), 400

    birth_date = str(data["birth_date"]).strip()
    if not re.fullmatch(r"\d{6}", birth_date):
        return jsonify({"error": "birth_date must be YYMMDD"}), 400

    emp_name = str(data["emp_name"]).strip()
    try:
        employee = Employee.query.filter_by(
            name=emp_name, birth_date=birth_date, is_active=True
        ).first()
    except OperationalError as exc:
        logger.error("Attendance create query failed: %s", exc)
        return _db_not_ready_json()

    if not employee:
        return jsonify({"error": "등록된 재직 직원만 입력할 수 있습니다."}), 403

    try:
        work_date = _parse_date(str(data["work_date"]).strip())
    except ValueError:
        return jsonify({"error": "work_date format must be YYYY-MM-DD"}), 400

    work_type = str(data.get("work_type", "normal")).strip()
    if work_type not in ALLOWED_WORK_TYPES:
        return jsonify({"error": f"Invalid work_type: {work_type}"}), 400

    try:
        exists = AttendanceRecord.query.filter_by(
            employee_id=employee.id,
            work_date=work_date,
        ).first()
    except OperationalError as exc:
        logger.error("Attendance duplicate-check query failed: %s", exc)
        return _db_not_ready_json()

    # 우선순위: admin > employee > excel
    # 관리자/직원 기록이 있으면 거부, 엑셀 기록이면 직원 입력으로 덮어쓰기
    if exists:
        if exists.source in ("admin", "employee"):
            return jsonify({"error": "같은 날짜의 근태 기록이 이미 존재합니다."}), 409
        # exists.source == "excel" → 직원 입력이 우선이므로 덮어씀

    clock_in = None
    clock_out = None
    total_hours = 0.0
    ot_hours = 0.0
    night_hours = 0.0
    holiday_hours = 0.0

    if work_type in TIME_REQUIRED_TYPES:
        clock_in = str(data.get("clock_in", "")).strip()
        clock_out = str(data.get("clock_out", "")).strip()
        if not (_validate_hhmm(clock_in) and _validate_hhmm(clock_out)):
            return jsonify({"error": "clock_in/clock_out format must be HH:MM"}), 400
        cfg = _get_cfg()
        try:
            day_type = _effective_day_type(work_date, cfg)
        except OperationalError as exc:
            logger.error("Attendance calendar lookup failed: %s", exc)
            return _db_not_ready_json()
        total_hours, ot_hours, night_hours, holiday_hours = calc_work_hours(
            clock_in,
            clock_out,
            cfg,
            work_date=work_date,
            calendar_day_type=day_type,
        )

    if exists and exists.source == "excel":
        # 엑셀 기록을 직원 입력으로 덮어씀
        exists.birth_date = employee.birth_date
        exists.emp_name = employee.name
        exists.dept = str(data.get("dept", "")).strip()
        exists.clock_in = clock_in
        exists.clock_out = clock_out
        exists.work_type = work_type
        exists.total_work_hours = total_hours
        exists.overtime_hours = ot_hours
        exists.night_hours = night_hours
        exists.holiday_work_hours = holiday_hours
        exists.source = "employee"
        try:
            db.session.commit()
            return jsonify({"success": True, "record": exists.to_dict()}), 200
        except Exception as exc:
            db.session.rollback()
            logger.error("Attendance overwrite error: %s", exc)
            return jsonify({"error": "서버 오류가 발생했습니다."}), 500

    record = AttendanceRecord(
        employee_id=employee.id,
        birth_date=employee.birth_date,
        emp_name=employee.name,
        dept=str(data.get("dept", "")).strip(),
        work_date=work_date,
        clock_in=clock_in,
        clock_out=clock_out,
        work_type=work_type,
        total_work_hours=total_hours,
        overtime_hours=ot_hours,
        night_hours=night_hours,
        holiday_work_hours=holiday_hours,
        source="employee",
    )

    try:
        db.session.add(record)
        db.session.commit()
        return jsonify({"success": True, "record": record.to_dict()}), 201
    except IntegrityError:
        db.session.rollback()
        return jsonify({"error": "같은 날짜의 근태 기록이 이미 존재합니다."}), 409
    except Exception as exc:
        db.session.rollback()
        logger.error("Attendance create error: %s", exc)
        return jsonify({"error": "서버 오류가 발생했습니다."}), 500


@attendance_bp.route("/api/attendance", methods=["GET"])
@require_admin
def list_attendance():
    query = AttendanceRecord.query

    employee_id = request.args.get("employee_id") or request.args.get("emp_id")
    if employee_id:
        if not str(employee_id).isdigit():
            return jsonify({"error": "employee_id must be integer"}), 400
        query = query.filter(AttendanceRecord.employee_id == int(employee_id))

    emp_name = request.args.get("emp_name", "").strip()
    if emp_name:
        query = query.filter(AttendanceRecord.emp_name.contains(emp_name))

    start = request.args.get("start_date")
    if start:
        try:
            query = query.filter(AttendanceRecord.work_date >= _parse_date(start))
        except ValueError:
            return jsonify({"error": "start_date format must be YYYY-MM-DD"}), 400

    end = request.args.get("end_date")
    if end:
        try:
            query = query.filter(AttendanceRecord.work_date <= _parse_date(end))
        except ValueError:
            return jsonify({"error": "end_date format must be YYYY-MM-DD"}), 400

    try:
        records = query.order_by(AttendanceRecord.work_date.asc()).all()
    except OperationalError as exc:
        logger.error("Attendance list query failed: %s", exc)
        return _db_not_ready_json()

    return jsonify({"records": [r.to_dict() for r in records]})


@attendance_bp.route("/attendance")
def attendance_page():
    today = date.today()
    try:
        today_records = (
            AttendanceRecord.query.filter(AttendanceRecord.work_date == today)
            .order_by(AttendanceRecord.emp_name.asc())
            .all()
        )
    except OperationalError as exc:
        logger.error("Attendance page query failed: %s", exc)
        return _db_not_ready_page()

    return render_template("attendance.html", records=today_records, today=today)


@attendance_bp.route("/admin/attendance-calendar")
@require_admin
def admin_attendance_calendar():
    month_text = request.args.get("month", date.today().strftime("%Y-%m"))
    if not _validate_month(month_text):
        month_text = date.today().strftime("%Y-%m")

    start_date, end_date = _month_bounds(month_text)
    try:
        overrides = (
            OperationCalendarDay.query.filter(
                OperationCalendarDay.work_date >= start_date,
                OperationCalendarDay.work_date <= end_date,
            )
            .order_by(OperationCalendarDay.work_date.asc())
            .all()
        )
    except OperationalError as exc:
        logger.error("Attendance calendar query failed: %s", exc)
        return _db_not_ready_page()

    override_map = {row.work_date: row for row in overrides}

    cfg = _get_cfg()
    days = []
    day = start_date.day
    while day <= end_date.day:
        current = date(start_date.year, start_date.month, day)
        default_type = _default_day_type(current, cfg)
        override = override_map.get(current)
        effective_type = override.day_type if override else default_type
        days.append(
            {
                "date": current,
                "default_type": default_type,
                "override_type": override.day_type if override else "default",
                "effective_type": effective_type,
                "note": override.note if override else "",
            }
        )
        day += 1

    return render_template(
        "admin_attendance_calendar.html",
        month=month_text,
        days=days,
    )


@attendance_bp.route("/admin/attendance-calendar", methods=["POST"])
@require_admin
def save_attendance_calendar():
    work_date_text = (request.form.get("work_date") or "").strip()
    day_type = (request.form.get("day_type") or "").strip()
    note = (request.form.get("note") or "").strip()[:200]
    month_text = (request.form.get("month") or "").strip()

    if not _validate_month(month_text):
        month_text = date.today().strftime("%Y-%m")

    try:
        work_date = _parse_date(work_date_text)
    except ValueError:
        return redirect(url_for("attendance.admin_attendance_calendar", month=month_text))

    try:
        row = OperationCalendarDay.query.filter_by(work_date=work_date).first()
    except OperationalError as exc:
        logger.error("Attendance calendar write query failed: %s", exc)
        return _db_not_ready_page()
    if day_type == "default":
        if row:
            db.session.delete(row)
            try:
                db.session.commit()
            except OperationalError as exc:
                db.session.rollback()
                logger.error("Attendance calendar delete commit failed: %s", exc)
                return _db_not_ready_page()
        return redirect(url_for("attendance.admin_attendance_calendar", month=month_text))

    if day_type not in CALENDAR_DAY_TYPES:
        return redirect(url_for("attendance.admin_attendance_calendar", month=month_text))

    if row:
        row.day_type = day_type
        row.note = note
    else:
        row = OperationCalendarDay(
            work_date=work_date,
            day_type=day_type,
            note=note,
        )
        db.session.add(row)
    try:
        db.session.commit()
    except OperationalError as exc:
        db.session.rollback()
        logger.error("Attendance calendar save commit failed: %s", exc)
        return _db_not_ready_page()

    return redirect(url_for("attendance.admin_attendance_calendar", month=month_text))


@attendance_bp.route("/admin/attendance")
@require_admin
def admin_attendance():
    start = request.args.get("start_date", "")
    end = request.args.get("end_date", "")
    emp_name = request.args.get("emp_name", "")
    work_type = request.args.get("work_type", "")

    query = AttendanceRecord.query
    if start:
        try:
            query = query.filter(AttendanceRecord.work_date >= _parse_date(start))
        except ValueError:
            start = ""
    if end:
        try:
            query = query.filter(AttendanceRecord.work_date <= _parse_date(end))
        except ValueError:
            end = ""
    if emp_name:
        query = query.filter(AttendanceRecord.emp_name.contains(emp_name))
    if work_type:
        query = query.filter(AttendanceRecord.work_type == work_type)

    page = request.args.get("page", 1, type=int)
    per_page = 50

    try:
        # 통계는 전체 필터 결과 기반
        from sqlalchemy import func as sa_func
        stats_query = query.with_entities(
            sa_func.count().label("total"),
            sa_func.sum(AttendanceRecord.total_work_hours).label("total_work"),
            sa_func.sum(AttendanceRecord.night_hours).label("total_night"),
            sa_func.sum(AttendanceRecord.overtime_hours).label("total_ot"),
            sa_func.sum(AttendanceRecord.holiday_work_hours).label("total_holiday"),
        ).first()

        day_count = query.filter(AttendanceRecord.work_type == "normal").count()
        night_count = query.filter(AttendanceRecord.work_type == "night").count()

        stats = {
            "total": stats_query.total or 0,
            "day": day_count,
            "night": night_count,
            "total_work": round(stats_query.total_work or 0, 1),
            "total_night": round(stats_query.total_night or 0, 1),
            "total_ot": round(stats_query.total_ot or 0, 1),
            "total_holiday": round(stats_query.total_holiday or 0, 1),
        }

        pagination = query.order_by(AttendanceRecord.work_date.desc()).paginate(
            page=page, per_page=per_page, error_out=False
        )
    except OperationalError as exc:
        logger.error("Admin attendance query failed: %s", exc)
        return _db_not_ready_page()

    today = date.today()

    # 고객사 매핑 (employee_id → site_name)
    emp_ids = list({r.employee_id for r in pagination.items})
    site_map = {}
    if emp_ids:
        rows = (
            db.session.query(Employee.id, Site.name)
            .outerjoin(Site, Employee.site_id == Site.id)
            .filter(Employee.id.in_(emp_ids))
            .all()
        )
        site_map = {eid: sname or "-" for eid, sname in rows}

    return render_template(
        "admin_attendance.html",
        records=pagination.items,
        pagination=pagination,
        start_date=start,
        end_date=end,
        emp_name=emp_name,
        work_type=work_type,
        stats=stats,
        today=today,
        site_map=site_map,
    )


@attendance_bp.route("/admin/attendance/excel")
@require_admin
def attendance_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font

    start = request.args.get("start_date", "")
    end = request.args.get("end_date", "")
    emp_name = request.args.get("emp_name", "")
    work_type = request.args.get("work_type", "")

    query = AttendanceRecord.query
    if start:
        try:
            query = query.filter(AttendanceRecord.work_date >= _parse_date(start))
        except ValueError:
            start = ""
    if end:
        try:
            query = query.filter(AttendanceRecord.work_date <= _parse_date(end))
        except ValueError:
            end = ""
    if emp_name:
        query = query.filter(AttendanceRecord.emp_name.contains(emp_name))
    if work_type:
        query = query.filter(AttendanceRecord.work_type == work_type)

    try:
        records = query.order_by(AttendanceRecord.work_date.asc()).all()
    except OperationalError as exc:
        logger.error("Attendance excel query failed: %s", exc)
        return _db_not_ready_page()

    wb = Workbook()
    ws = wb.active
    ws.title = "근태기록"

    headers = [
        "직원ID",
        "이름",
        "부서",
        "날짜",
        "출근",
        "퇴근",
        "구분",
        "총근무(h)",
        "잔업(h)",
        "야간(h)",
        "휴일(h)",
        "출처",
    ]
    bold = Font(bold=True)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = bold

    work_type_labels = {
        "normal": "주간",
        "night": "야간",
        "annual": "연차",
        "absent": "결근",
        "holiday": "휴무",
        "early": "조퇴",
    }

    for i, rec in enumerate(records, 2):
        ws.cell(row=i, column=1, value=rec.employee_id)
        ws.cell(row=i, column=2, value=rec.emp_name)
        ws.cell(row=i, column=3, value=rec.dept)
        ws.cell(
            row=i,
            column=4,
            value=rec.work_date.strftime("%Y-%m-%d") if rec.work_date else "",
        )
        ws.cell(row=i, column=5, value=rec.clock_in)
        ws.cell(row=i, column=6, value=rec.clock_out)
        ws.cell(row=i, column=7, value=work_type_labels.get(rec.work_type, rec.work_type))
        ws.cell(row=i, column=8, value=rec.total_work_hours)
        ws.cell(row=i, column=9, value=rec.overtime_hours)
        ws.cell(row=i, column=10, value=rec.night_hours)
        ws.cell(row=i, column=11, value=rec.holiday_work_hours or 0)
        source_labels = {"employee": "직원", "excel": "엑셀", "admin": "관리자"}
        ws.cell(row=i, column=12, value=source_labels.get(rec.source, rec.source or "직원"))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"근태기록_{start or 'all'}_{end or 'all'}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@attendance_bp.route("/api/attendance/admin", methods=["POST"])
@require_admin
def admin_create_attendance():
    """관리자가 수동으로 근태 기록을 추가한다.

    충돌 정책:
    - 같은 직원+날짜에 이미 기록이 있으면 덮어쓰기(overwrite) 옵션 지원.
    - overwrite=true 이면 기존 기록을 갱신, false(기본)이면 409 반환.
    """
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "JSON body required"}), 400

    emp_name = str(data.get("emp_name", "")).strip()
    birth_date = str(data.get("birth_date", "")).strip()
    if not emp_name or not birth_date:
        return jsonify({"error": "이름과 생년월일은 필수입니다."}), 400
    if not re.fullmatch(r"\d{6}", birth_date):
        return jsonify({"error": "생년월일은 YYMMDD 6자리 숫자여야 합니다."}), 400

    employee = Employee.query.filter_by(
        name=emp_name, birth_date=birth_date, is_active=True
    ).first()
    if not employee:
        return jsonify({"error": "등록된 재직 직원을 찾을 수 없습니다."}), 404

    try:
        work_date = _parse_date(str(data.get("work_date", "")).strip())
    except ValueError:
        return jsonify({"error": "날짜 형식이 올바르지 않습니다. (YYYY-MM-DD)"}), 400

    work_type = str(data.get("work_type", "normal")).strip()
    if work_type not in ALLOWED_WORK_TYPES:
        return jsonify({"error": f"잘못된 근무 유형: {work_type}"}), 400

    clock_in = str(data.get("clock_in", "")).strip() or None
    clock_out = str(data.get("clock_out", "")).strip() or None

    total_hours = 0.0
    ot_hours = 0.0
    night_h = 0.0
    holiday_h = 0.0

    if work_type in TIME_REQUIRED_TYPES:
        if not (_validate_hhmm(clock_in or "") and _validate_hhmm(clock_out or "")):
            return jsonify({"error": "출퇴근 시간을 정확히 입력해주세요. (HH:MM)"}), 400
        cfg = _get_cfg()
        try:
            day_type = _effective_day_type(work_date, cfg)
        except OperationalError:
            return _db_not_ready_json()
        total_hours, ot_hours, night_h, holiday_h = calc_work_hours(
            clock_in, clock_out, cfg,
            work_date=work_date,
            calendar_day_type=day_type,
        )

    # 충돌 확인
    existing = AttendanceRecord.query.filter_by(
        employee_id=employee.id, work_date=work_date
    ).first()

    overwrite = str(data.get("overwrite", "")).lower() in ("true", "1", "yes")

    if existing and not overwrite:
        source_label = {"employee": "직원 입력", "excel": "엑셀 업로드", "admin": "관리자 입력"}
        return jsonify({
            "error": "같은 날짜의 근태 기록이 이미 존재합니다.",
            "conflict": True,
            "existing_source": source_label.get(existing.source, existing.source),
            "existing_id": existing.id,
        }), 409

    try:
        if existing and overwrite:
            existing.emp_name = employee.name
            existing.birth_date = employee.birth_date
            existing.dept = str(data.get("dept", existing.dept or "")).strip()
            existing.clock_in = clock_in
            existing.clock_out = clock_out
            existing.work_type = work_type
            existing.total_work_hours = total_hours
            existing.overtime_hours = ot_hours
            existing.night_hours = night_h
            existing.holiday_work_hours = holiday_h
            existing.source = "admin"
            db.session.commit()
            return jsonify({"success": True, "record": existing.to_dict(), "overwritten": True})
        else:
            record = AttendanceRecord(
                employee_id=employee.id,
                birth_date=employee.birth_date,
                emp_name=employee.name,
                dept=str(data.get("dept", "")).strip(),
                work_date=work_date,
                clock_in=clock_in,
                clock_out=clock_out,
                work_type=work_type,
                total_work_hours=total_hours,
                overtime_hours=ot_hours,
                night_hours=night_h,
                holiday_work_hours=holiday_h,
                source="admin",
            )
            db.session.add(record)
            db.session.commit()
            return jsonify({"success": True, "record": record.to_dict()}), 201
    except IntegrityError:
        db.session.rollback()
        return jsonify({"error": "중복 근태 기록입니다."}), 409
    except Exception as exc:
        db.session.rollback()
        logger.error("Admin attendance create error: %s", exc)
        return jsonify({"error": "서버 오류가 발생했습니다."}), 500


@attendance_bp.route("/api/attendance/<int:record_id>", methods=["PUT"])
@require_admin
def update_attendance(record_id):
    record = db.get_or_404(AttendanceRecord, record_id)
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "JSON body required"}), 400

    if "employee_id" in data:
        employee_id = str(data.get("employee_id", "")).strip()
        if not employee_id.isdigit():
            return jsonify({"error": "employee_id must be integer"}), 400
        employee = db.session.get(Employee, int(employee_id))
        if not employee:
            return jsonify({"error": "employee not found"}), 404
        record.employee_id = employee.id
        record.emp_name = employee.name
        record.birth_date = employee.birth_date

    if "emp_name" in data or "birth_date" in data:
        name = str(data.get("emp_name", record.emp_name)).strip()
        birth = str(data.get("birth_date", record.birth_date)).strip()
        employee = Employee.query.filter_by(name=name, birth_date=birth).first()
        if employee:
            record.employee_id = employee.id
            record.emp_name = employee.name
            record.birth_date = employee.birth_date

    if "dept" in data:
        record.dept = str(data.get("dept", "")).strip()

    if "work_date" in data:
        try:
            record.work_date = _parse_date(str(data["work_date"]).strip())
        except ValueError:
            return jsonify({"error": "work_date format must be YYYY-MM-DD"}), 400

    if "work_type" in data:
        work_type = str(data["work_type"]).strip()
        if work_type not in ALLOWED_WORK_TYPES:
            return jsonify({"error": f"Invalid work_type: {work_type}"}), 400
        record.work_type = work_type

    if "clock_in" in data:
        val = str(data.get("clock_in") or "").strip()
        record.clock_in = val or None
    if "clock_out" in data:
        val = str(data.get("clock_out") or "").strip()
        record.clock_out = val or None

    duplicate = AttendanceRecord.query.filter(
        AttendanceRecord.employee_id == record.employee_id,
        AttendanceRecord.work_date == record.work_date,
        AttendanceRecord.id != record.id,
    ).first()
    if duplicate:
        return jsonify({"error": "같은 날짜의 근태 기록이 이미 존재합니다."}), 409

    if record.work_type in TIME_REQUIRED_TYPES:
        if not (_validate_hhmm(record.clock_in or "") and _validate_hhmm(record.clock_out or "")):
            return jsonify({"error": "clock_in/clock_out format must be HH:MM"}), 400
        cfg = _get_cfg()
        try:
            day_type = _effective_day_type(record.work_date, cfg)
        except OperationalError as exc:
            logger.error("Attendance calendar lookup failed during update: %s", exc)
            return _db_not_ready_json()
        total, ot, night, holiday = calc_work_hours(
            record.clock_in,
            record.clock_out,
            cfg,
            work_date=record.work_date,
            calendar_day_type=day_type,
        )
        record.total_work_hours = total
        record.overtime_hours = ot
        record.night_hours = night
        record.holiday_work_hours = holiday
    else:
        record.clock_in = None
        record.clock_out = None
        record.total_work_hours = 0.0
        record.overtime_hours = 0.0
        record.night_hours = 0.0
        record.holiday_work_hours = 0.0

    try:
        db.session.commit()
        return jsonify({"success": True, "record": record.to_dict()})
    except IntegrityError:
        db.session.rollback()
        return jsonify({"error": "중복 근태 기록입니다."}), 409
    except Exception as exc:
        db.session.rollback()
        logger.error("Attendance update error: %s", exc)
        return jsonify({"error": "서버 오류가 발생했습니다."}), 500


@attendance_bp.route("/api/attendance/<int:record_id>", methods=["DELETE"])
@require_admin
def delete_attendance(record_id):
    record = db.get_or_404(AttendanceRecord, record_id)
    try:
        db.session.delete(record)
        db.session.commit()
        return jsonify({"success": True})
    except Exception as exc:
        db.session.rollback()
        logger.error("Attendance delete error: %s", exc)
        return jsonify({"error": "DELETE failed"}), 500


@attendance_bp.route("/api/attendance/bulk-delete", methods=["POST"])
@require_admin
def bulk_delete_attendance():
    """선택 삭제 또는 필터 조건 전체 삭제."""
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "JSON body required"}), 400

    try:
        # 방법 1: ID 목록으로 삭제
        ids = data.get("ids")
        if ids and isinstance(ids, list):
            ids = [int(i) for i in ids]
            deleted = AttendanceRecord.query.filter(AttendanceRecord.id.in_(ids)).delete(
                synchronize_session=False
            )
            db.session.commit()
            return jsonify({"success": True, "deleted": deleted})

        # 방법 2: 필터 조건으로 전체 삭제
        filt = data.get("filter", {})
        query = AttendanceRecord.query
        start = filt.get("start_date", "")
        end = filt.get("end_date", "")
        emp_name = filt.get("emp_name", "")
        work_type = filt.get("work_type", "")

        # 최소 하나의 필터 조건 필수 (전체 삭제 방지)
        if not any([start, end, emp_name, work_type]):
            return jsonify({"error": "최소 하나의 필터 조건을 지정해야 합니다."}), 400

        if start:
            try:
                query = query.filter(AttendanceRecord.work_date >= _parse_date(start))
            except ValueError:
                pass
        if end:
            try:
                query = query.filter(AttendanceRecord.work_date <= _parse_date(end))
            except ValueError:
                pass
        if emp_name:
            query = query.filter(AttendanceRecord.emp_name.contains(emp_name))
        if work_type:
            query = query.filter(AttendanceRecord.work_type == work_type)

        deleted = query.delete(synchronize_session=False)
        db.session.commit()
        return jsonify({"success": True, "deleted": deleted})

    except Exception as exc:
        db.session.rollback()
        logger.error("Bulk delete error: %s", exc)
        return jsonify({"error": "삭제 처리 중 오류가 발생했습니다."}), 500


# ── 근태 엑셀 업로드 ──────────────────────────────────

@attendance_bp.route("/admin/attendance/import", methods=["GET", "POST"])
@require_admin
def import_attendance():
    """GET: 업로드 폼, POST: 미리보기(dry-run)"""
    if request.method == "GET":
        return render_template("admin_attendance_import.html", result=None)

    file = request.files.get("file")
    if not file or not file.filename:
        return render_template(
            "admin_attendance_import.html",
            result={"errors": ["파일을 선택해주세요."]},
        )

    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in (".xlsx", ".xls"):
        return render_template(
            "admin_attendance_import.html",
            result={"errors": ["엑셀 파일(.xlsx, .xls)만 업로드 가능합니다."]},
        )

    from services.attendance_import import import_attendance_to_db, parse_attendance_excel

    try:
        parsed = parse_attendance_excel(file, filename=file.filename)
    except Exception as exc:
        logger.error("엑셀 파싱 오류: %s", exc)
        return render_template(
            "admin_attendance_import.html",
            result={"errors": [f"엑셀 파싱 실패: {exc}"]},
        )

    if parsed.get("errors"):
        return render_template(
            "admin_attendance_import.html",
            result={"errors": parsed["errors"]},
        )

    # dry_run으로 미리보기 결과 생성
    preview = import_attendance_to_db(parsed, dry_run=True)
    preview["month_str"] = parsed["month_str"]
    preview["site_name"] = parsed["site_name"]
    preview["employee_count"] = len(parsed["employees"])

    # 파싱 결과를 임시 파일에 저장 (세션 쿠키 4KB 제한 회피)
    import json
    import uuid

    from flask import session as flask_session

    serializable = _serialize_parsed(parsed)
    import_id = uuid.uuid4().hex[:12]
    tmp_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "tmp")
    os.makedirs(tmp_dir, exist_ok=True)
    tmp_path = os.path.join(tmp_dir, f"import_{import_id}.json")
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(serializable, f, ensure_ascii=False)
    flask_session["attendance_import_id"] = import_id

    return render_template(
        "admin_attendance_import.html",
        result=preview,
        preview=True,
    )


@attendance_bp.route("/admin/attendance/import/execute", methods=["POST"])
@require_admin
def execute_import():
    """미리보기 확인 후 실제 DB 저장 실행."""
    import json

    from flask import flash, session as flask_session

    from services.attendance_import import import_attendance_to_db

    import_id = flask_session.pop("attendance_import_id", None)
    if not import_id:
        flash("업로드 데이터가 만료되었습니다. 다시 업로드해주세요.", "error")
        return redirect(url_for("attendance.import_attendance"))

    tmp_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "tmp")
    tmp_path = os.path.join(tmp_dir, f"import_{import_id}.json")

    if not os.path.exists(tmp_path):
        flash("업로드 데이터가 만료되었습니다. 다시 업로드해주세요.", "error")
        return redirect(url_for("attendance.import_attendance"))

    try:
        with open(tmp_path, "r", encoding="utf-8") as f:
            parsed = json.load(f)
        parsed = _deserialize_parsed(parsed)
        os.remove(tmp_path)  # 사용 후 삭제
    except Exception as exc:
        logger.error("Import 데이터 복원 실패: %s", exc)
        flash("데이터 복원 실패. 다시 업로드해주세요.", "error")
        return redirect(url_for("attendance.import_attendance"))

    result = import_attendance_to_db(parsed, dry_run=False)
    result["month_str"] = parsed["month_str"]
    result["site_name"] = parsed["site_name"]
    result["employee_count"] = len(parsed["employees"])

    from flask import flash

    if not result["errors"]:
        flash(
            f"{result['month_str']} {result['site_name']} 근태 업로드 완료: "
            f"신규 {result['created']}건, 갱신 {result['updated']}건",
            "success",
        )
    else:
        flash("일부 오류가 발생했습니다. 아래 결과를 확인해주세요.", "error")

    return render_template(
        "admin_attendance_import.html",
        result=result,
        executed=True,
    )


def _serialize_parsed(parsed: dict) -> dict:
    """파싱 결과를 JSON 직렬화 가능하게 변환."""
    import copy

    data = copy.deepcopy(parsed)
    for emp in data.get("employees", []):
        if emp.get("hire_date"):
            emp["hire_date"] = emp["hire_date"].isoformat()
        if emp.get("resign_date"):
            emp["resign_date"] = emp["resign_date"].isoformat()
        new_days = {}
        for d, v in emp.get("days", {}).items():
            key = d.isoformat() if isinstance(d, date) else str(d)
            new_days[key] = v
        emp["days"] = new_days
    return data


def _deserialize_parsed(data: dict) -> dict:
    """JSON에서 복원된 파싱 결과를 원래 타입으로 변환."""
    for emp in data.get("employees", []):
        if emp.get("hire_date"):
            emp["hire_date"] = date.fromisoformat(emp["hire_date"])
        if emp.get("resign_date"):
            emp["resign_date"] = date.fromisoformat(emp["resign_date"])
        new_days = {}
        for d_str, v in emp.get("days", {}).items():
            new_days[date.fromisoformat(d_str)] = v
        emp["days"] = new_days
    return data
