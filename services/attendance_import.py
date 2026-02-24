"""근태 엑셀 파일 파싱 및 DB 저장 서비스."""

import logging
import os
import re
from datetime import date, datetime

from models import AttendanceRecord, Employee, db

logger = logging.getLogger(__name__)

# 엑셀 구조 상수
_HEADER_ROW = 5           # 날짜 헤더 행
_DATA_START_ROW = 8       # 첫 직원 데이터 시작 행
_ROWS_PER_EMPLOYEE = 6    # 직원당 행 수
_DATE_COL_START = 8       # H열 = 1일
_DATE_COL_END = 38        # AL열 = 31일
_SUMMARY_COL = 39         # AM열 = 합계

# 6행 구분 (G열 값 → 내부 키)
_CATEGORY_MAP = {
    "기본": "base",
    "연장": "ot",
    "심야": "night",
    "특근": "holiday",
    "특연": "holiday_ot",
    "지조": "late",
}


def _safe_float(value) -> float:
    """셀 값을 float로 안전 변환. 문자열('연차' 등)이면 0."""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0.0


def _to_date(value) -> date | None:
    """엑셀 셀 값을 date로 변환."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _extract_site_name(filename: str) -> str:
    """파일명에서 업체명 추출. 예: 'Humetix - 영진팩 1월 근태.xlsx' → '영진팩'"""
    name = os.path.splitext(os.path.basename(filename))[0]
    # 'Humetix - 영진팩 1월 근태' 패턴
    m = re.search(r"[-–]\s*(.+?)\s*\d+월", name)
    if m:
        return m.group(1).strip()
    # '-' 뒤 첫 단어
    m = re.search(r"[-–]\s*(\S+)", name)
    if m:
        return m.group(1).strip()
    return name


def parse_attendance_excel(file_stream, filename: str = "") -> dict:
    """
    근태 엑셀 파일을 파싱하여 구조화된 데이터를 반환.

    Args:
        file_stream: 엑셀 파일 스트림 (BytesIO 또는 FileStorage)
        filename: 원본 파일명 (업체명 추출용)

    Returns:
        {
            "year": 2026,
            "month": 1,
            "month_str": "2026-01",
            "site_name": "영진팩",
            "employees": [
                {
                    "name": "이은비",
                    "hire_date": date(2022, 11, 10),
                    "resign_date": None,
                    "hourly_wage": 10320,
                    "days": {
                        date(2026,1,2): {
                            "base": 8, "ot": 0, "night": 0,
                            "holiday": 0, "holiday_ot": 0, "late": 0,
                            "is_annual": False
                        },
                        ...
                    },
                    "summary": {"base": 209, "ot": 30, "night": 0, ...}
                },
                ...
            ],
            "errors": []
        }
    """
    import openpyxl

    wb = openpyxl.load_workbook(file_stream, data_only=True, read_only=True)
    ws = wb.worksheets[0]

    errors = []

    # ── 연월 추출 ──
    year_val = None
    month_val = None
    for row in ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=20, values_only=False):
        for cell in row:
            if cell.column == 10:  # J열
                year_val = cell.value
            if cell.column == 13:  # M열
                month_val = cell.value

    if year_val is None or month_val is None:
        # 시트명에서 추출 시도: "1월 근태"
        sheet_name = ws.title or ""
        m = re.search(r"(\d+)월", sheet_name)
        if m and month_val is None:
            month_val = int(m.group(1))
        if year_val is None:
            year_val = date.today().year

    year = int(year_val)
    month = int(month_val)
    month_str = f"{year}-{month:02d}"

    # ── 업체명 추출 ──
    site_name = ""
    # E2:F3 merged cell
    for row in ws.iter_rows(min_row=2, max_row=2, min_col=5, max_col=5, values_only=False):
        for cell in row:
            if cell.value:
                site_name = str(cell.value).strip()
    if not site_name and filename:
        site_name = _extract_site_name(filename)

    # ── 날짜 헤더 추출 (Row 5, H~AL) ──
    date_columns = {}  # col_index → date
    for row in ws.iter_rows(min_row=_HEADER_ROW, max_row=_HEADER_ROW,
                            min_col=_DATE_COL_START, max_col=_DATE_COL_END,
                            values_only=False):
        for cell in row:
            d = _to_date(cell.value)
            if d:
                date_columns[cell.column] = d

    if not date_columns:
        wb.close()
        return {
            "year": year, "month": month, "month_str": month_str,
            "site_name": site_name, "employees": [],
            "errors": ["날짜 헤더를 찾을 수 없습니다 (Row 5, H~AL열)."],
        }

    # ── 직원 데이터 파싱 ──
    employees = []

    # read_only 모드에서는 전체 행을 미리 읽어둔다
    all_rows = {}
    for row in ws.iter_rows(min_row=_DATA_START_ROW, max_row=ws.max_row or 500,
                            min_col=2, max_col=_SUMMARY_COL + 1,
                            values_only=False):
        for cell in row:
            if cell.row not in all_rows:
                all_rows[cell.row] = {}
            all_rows[cell.row][cell.column] = cell.value

    wb.close()

    row_num = _DATA_START_ROW
    while row_num in all_rows:
        row_data = all_rows.get(row_num, {})
        g_val = row_data.get(7)  # G열 = 구분

        if g_val != "기본":
            row_num += 1
            continue

        # 첫 행(기본): 연번(B), 성명(C), 입사일(D), 퇴사일(E), 시급(F)
        seq_num = row_data.get(2)  # B열
        emp_name = row_data.get(3)  # C열
        hire_date_raw = row_data.get(4)  # D열
        resign_date_raw = row_data.get(5)  # E열
        hourly_wage_raw = row_data.get(6)  # F열

        # 이름 없는 빈 블록 → 스킵
        if not emp_name:
            row_num += _ROWS_PER_EMPLOYEE
            continue

        # 이름 정리: 줄바꿈 제거, 괄호 내 부서명 제거, 공백 정리
        emp_name = str(emp_name).strip()
        emp_name = re.sub(r"[\r\n]+", " ", emp_name)  # 줄바꿈 → 공백
        emp_name = re.sub(r"\s*\(.*?\)", "", emp_name)  # (부서명) 제거
        emp_name = re.sub(r"\s{2,}", " ", emp_name).strip()  # 다중 공백 정리
        hire_date = _to_date(hire_date_raw)
        resign_date = _to_date(resign_date_raw)
        hourly_wage = _safe_float(hourly_wage_raw)

        # 6행 데이터 수집
        category_data = {}  # "base" → {col: value, ...}
        annual_dates = set()  # 연차인 날짜

        for offset in range(_ROWS_PER_EMPLOYEE):
            r = row_num + offset
            r_data = all_rows.get(r, {})
            g = r_data.get(7, "")
            cat_key = _CATEGORY_MAP.get(g)
            if not cat_key:
                continue

            col_values = {}
            for col_idx in range(_DATE_COL_START, _DATE_COL_END + 1):
                val = r_data.get(col_idx)
                if val is None:
                    continue
                # "연차" 텍스트 감지 (연장 행에서)
                if isinstance(val, str) and "연차" in val:
                    if col_idx in date_columns:
                        annual_dates.add(date_columns[col_idx])
                    continue
                num = _safe_float(val)
                if num != 0:
                    col_values[col_idx] = num

            # 합계(AM열)
            summary_val = _safe_float(r_data.get(_SUMMARY_COL))
            category_data[cat_key] = {
                "values": col_values,
                "summary": summary_val,
            }

        # 날짜별 데이터 구성
        days = {}
        for col_idx, work_date in date_columns.items():
            day_data = {
                "base": 0.0, "ot": 0.0, "night": 0.0,
                "holiday": 0.0, "holiday_ot": 0.0, "late": 0.0,
                "is_annual": work_date in annual_dates,
            }
            has_data = work_date in annual_dates
            for cat_key, cat_info in category_data.items():
                v = cat_info["values"].get(col_idx, 0.0)
                if v != 0:
                    day_data[cat_key] = v
                    has_data = True

            if has_data:
                days[work_date] = day_data

        # 요약
        summary = {}
        for cat_key in ("base", "ot", "night", "holiday", "holiday_ot", "late"):
            summary[cat_key] = category_data.get(cat_key, {}).get("summary", 0.0)
        summary["annual"] = len(annual_dates)

        employees.append({
            "name": emp_name,
            "hire_date": hire_date,
            "resign_date": resign_date,
            "hourly_wage": hourly_wage,
            "days": days,
            "summary": summary,
        })

        row_num += _ROWS_PER_EMPLOYEE

    return {
        "year": year,
        "month": month,
        "month_str": month_str,
        "site_name": site_name,
        "employees": employees,
        "errors": errors,
    }


def import_attendance_to_db(parsed_data: dict, dry_run: bool = False) -> dict:
    """
    파싱된 근태 데이터를 DB에 저장.

    Args:
        parsed_data: parse_attendance_excel() 반환값
        dry_run: True이면 실제 저장 없이 결과만 반환

    Returns:
        {
            "created": int,
            "updated": int,
            "skipped": int,
            "new_employees": [{"name": ..., "hire_date": ...}],
            "matched_employees": [{"name": ..., "id": ...}],
            "errors": [...],
            "warnings": [...],
            "total_records": int,
        }
    """
    result = {
        "created": 0,
        "updated": 0,
        "skipped": 0,
        "new_employees": [],
        "matched_employees": [],
        "errors": list(parsed_data.get("errors", [])),
        "warnings": [],
        "total_records": 0,
    }

    employees_data = parsed_data.get("employees", [])
    if not employees_data:
        result["errors"].append("파싱된 직원 데이터가 없습니다.")
        return result

    # ── 직원 매칭 ──
    emp_map = {}  # 엑셀 이름 → Employee 객체
    for emp_data in employees_data:
        name = emp_data["name"]
        if name in emp_map:
            continue  # 이미 매칭됨

        matches = Employee.query.filter_by(name=name, is_active=True).all()
        if len(matches) == 1:
            emp_map[name] = matches[0]
            result["matched_employees"].append({"name": name, "id": matches[0].id})
        elif len(matches) == 0:
            # 자동 등록
            if dry_run:
                emp_map[name] = None  # dry_run에서는 None 표시
                result["new_employees"].append({
                    "name": name,
                    "hire_date": emp_data.get("hire_date"),
                })
            else:
                new_emp = Employee(
                    name=name,
                    birth_date="000000",
                    hire_date=emp_data.get("hire_date"),
                    is_active=True,
                )
                db.session.add(new_emp)
                db.session.flush()  # ID 할당
                emp_map[name] = new_emp
                result["new_employees"].append({
                    "name": name,
                    "hire_date": emp_data.get("hire_date"),
                    "id": new_emp.id,
                })
                logger.info("자동 등록: %s (hire_date=%s)", name, emp_data.get("hire_date"))
        else:
            result["errors"].append(
                f"동명이인: {name} ({len(matches)}명) — 직원관리에서 확인 필요"
            )
            continue

    # ── 근태 레코드 생성/갱신 ──
    for emp_data in employees_data:
        name = emp_data["name"]
        employee = emp_map.get(name)

        if employee is None and not dry_run:
            continue
        if name not in emp_map:
            continue  # 동명이인 등으로 매칭 실패

        # 합계 검증 (2시간 이상 차이만 경고 — 엑셀 수식이 수동값이므로 소폭 차이는 허용)
        for cat_key, label in [("base", "기본"), ("ot", "연장"), ("night", "심야"),
                                ("holiday", "특근"), ("holiday_ot", "특연"), ("late", "지조")]:
            excel_sum = emp_data["summary"].get(cat_key, 0)
            calc_sum = sum(
                day.get(cat_key, 0) for day in emp_data["days"].values()
            )
            if abs(excel_sum - calc_sum) > 2.0:
                result["warnings"].append(
                    f"합계 불일치: {name} {label} 엑셀={excel_sum} 계산={calc_sum}"
                )

        for work_date, day_data in emp_data["days"].items():
            result["total_records"] += 1

            base = day_data.get("base", 0)
            ot = day_data.get("ot", 0)
            night = day_data.get("night", 0)
            holiday = day_data.get("holiday", 0)
            holiday_ot = day_data.get("holiday_ot", 0)
            late = day_data.get("late", 0)
            is_annual = day_data.get("is_annual", False)

            total_hours = base + ot + night + holiday + holiday_ot - late
            overtime = ot + holiday_ot

            # work_type 결정
            if is_annual:
                work_type = "annual"
            elif base == 0 and holiday > 0:
                work_type = "holiday"
            elif night > 0:
                work_type = "night"
            else:
                work_type = "normal"

            if dry_run:
                result["created"] += 1
                continue

            # upsert: 기존 레코드 있으면 UPDATE
            existing = AttendanceRecord.query.filter_by(
                employee_id=employee.id,
                work_date=work_date,
            ).first()

            if existing:
                # 우선순위: admin > employee > excel
                # 상위 source의 기록은 엑셀로 덮어쓰지 않음
                if existing.source in ("admin", "employee"):
                    result["skipped"] += 1
                    result["warnings"].append(
                        f"건너뜀: {name} {work_date} — 기존 "
                        f"{'관리자' if existing.source == 'admin' else '직원'} "
                        f"입력 기록 보호"
                    )
                    continue
                existing.total_work_hours = total_hours
                existing.overtime_hours = overtime
                existing.night_hours = night
                existing.holiday_work_hours = holiday
                existing.work_type = work_type
                existing.source = "excel"
                result["updated"] += 1
            else:
                record = AttendanceRecord(
                    employee_id=employee.id,
                    birth_date=employee.birth_date,
                    emp_name=employee.name,
                    dept="",
                    work_date=work_date,
                    work_type=work_type,
                    total_work_hours=total_hours,
                    overtime_hours=overtime,
                    night_hours=night,
                    holiday_work_hours=holiday,
                    source="excel",
                )
                db.session.add(record)
                result["created"] += 1

    if not dry_run:
        try:
            db.session.commit()
        except Exception as exc:
            db.session.rollback()
            logger.error("근태 import 커밋 실패: %s", exc)
            result["errors"].append(f"DB 저장 실패: {exc}")

    if result["new_employees"]:
        result["warnings"].append(
            f"신규 직원 {len(result['new_employees'])}명 자동 등록 "
            f"(생년월일 미설정 → 직원관리에서 수정 필요)"
        )

    return result
