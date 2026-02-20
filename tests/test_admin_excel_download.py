from io import BytesIO

from openpyxl import load_workbook

from models import Application, db


def _login_as_admin(client):
    with client.session_transaction() as sess:
        sess["is_admin"] = True


def test_download_excel_with_selected_columns(client):
    _login_as_admin(client)

    with client.application.app_context():
        db.session.add(
            Application(
                id="app-1",
                name="홍길동",
                phone="010-1234-5678",
                email="hong@example.com",
                address="서울",
                status="new",
            )
        )
        db.session.commit()

    response = client.get("/download_excel?excel_columns=name,phone,status")

    assert response.status_code == 200
    assert response.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    workbook = load_workbook(BytesIO(response.data))
    worksheet = workbook.active

    headers = [worksheet.cell(row=1, column=i).value for i in range(1, 4)]
    values = [worksheet.cell(row=2, column=i).value for i in range(1, 4)]

    assert headers == ["이름", "연락처", "상태"]
    assert values == ["홍길동", "010-1234-5678", "신규"]
