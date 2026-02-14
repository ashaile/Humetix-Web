import os
import json
import uuid
from flask import Flask, render_template, request, redirect, url_for, session, send_file, send_from_directory
from werkzeug.utils import secure_filename
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, Alignment, Border, Side

app = Flask(__name__)
app.secret_key = 'humetix_secure_key'  # 보안을 위해 실제 배포시 변경 권장

# 설정
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# 업로드 폴더 절대경로
UPLOAD_DIR = os.path.join(BASE_DIR, 'uploads')
# JSON 데이터 파일 절대경로
DATA_FILE = os.path.join(BASE_DIR, 'applications.json') 
ADMIN_PASSWORD = "3326"

if not os.path.exists(UPLOAD_DIR):
    os.makedirs(UPLOAD_DIR)

# JSON 데이터 로드/저장 헬퍼 함수
def load_data():
    if not os.path.exists(DATA_FILE):
        return []
    try:
        with open(DATA_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except:
        return []

def save_data(data):
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=4)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/apply')
def apply():
    return render_template('apply.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        password = request.form.get('password')
        if password == ADMIN_PASSWORD:
            session['is_admin'] = True
            return redirect(url_for('master_view'))
        else:
            return "<script>alert('비밀번호가 틀렸습니다.'); history.back();</script>"
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('is_admin', None)
    return redirect(url_for('index'))

@app.route('/submit', methods=['POST'])
def submit():
    try:
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        file_now = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # 1. 신분증 사진 처리
        id_card = request.files.get('id_card')
        photo_filename = ""
        
        if id_card and id_card.filename != '':
            ext = os.path.splitext(id_card.filename)[1]
            # 안전한 파일명 생성
            photo_name = f"{file_now}_id{ext}"
            photo_path = os.path.join(UPLOAD_DIR, photo_name)
            id_card.save(photo_path)
            photo_filename = photo_name

        # 2. 데이터 수집 (JSON 구조)
        application_id = str(uuid.uuid4()) # 고유 ID 생성
        new_entry = {
            "id": application_id,
            "timestamp": now,
            "photo": photo_filename,
            "info": {
                "name": request.form.get('name'),
                "birth": request.form.get('birth'),
                "phone": request.form.get('phone'),
                "email": request.form.get('email'),
                "address": request.form.get('address'),
            },
            "career": [],
            "body": {
                "height": request.form.get('height'),
                "weight": request.form.get('weight'),
                "vision": f"{request.form.get('vision_type')} {request.form.get('vision_value')}",
                "shoes": request.form.get('shoes'),
                "tshirt": request.form.get('tshirt'),
            },
            "work_condition": {
                "shift": request.form.get('shift'),
                "posture": request.form.get('posture'),
                "overtime": request.form.get('overtime'),
                "holiday": request.form.get('holiday'),
                "interview_date": request.form.get('interview_date'),
                "start_date": request.form.get('start_date'),
                "agree": request.form.get('agree')
            }
        }

        # 경력사항 루프 처리
        for i in range(1, 4):
            company = request.form.get(f'company{i}')
            if company:
                new_entry["career"].append({
                    "company": company,
                    "start": request.form.get(f'exp_start{i}'),
                    "end": request.form.get(f'exp_end{i}'),
                    "role": request.form.get(f'job_role{i}'),
                    "reason": request.form.get(f'reason{i}')
                })

        # 3. JSON 저장
        data = load_data()
        data.append(new_entry)
        save_data(data)
        
        return "<h1>지원서 접수 완료!</h1><script>setTimeout(function(){location.href='/';}, 2000);</script>"
        
    except Exception as e:
        return f"<h1>오류 발생: {str(e)}</h1>"

@app.route('/humetix_master_99')
def master_view():
    if not session.get('is_admin'):
        return redirect(url_for('login'))
    
    data = load_data()
    # 최신순 정렬
    data.reverse()
    
    # HTML 생성
    html = '''
    <style>
        body { font-family: 'Noto Sans KR', sans-serif; background: #f4f7f9; padding: 20px; }
        .container { max-width: 1000px; margin: 0 auto; background: white; padding: 30px; border-radius: 15px; box-shadow: 0 5px 15px rgba(0,0,0,0.1); }
        .header { display: flex; justify-content: space-between; align-items: center; margin-bottom: 30px; padding-bottom: 20px; border-bottom: 2px solid #003057; }
        .btn { padding: 8px 15px; border: none; border-radius: 5px; cursor: pointer; font-weight: bold; text-decoration: none; display: inline-block;}
        .btn-excel { background: #28a745; color: white; }
        .btn-delete { background: #dc3545; color: white; }
        .btn-logout { background: #6c757d; color: white; }
        .card { border: 1px solid #ddd; padding: 20px; margin-bottom: 20px; border-radius: 10px; background: #fff; position: relative; }
        .card-header { display: flex; justify-content: space-between; align-items: center; background: #f8f9fa; padding: 10px; margin: -20px -20px 15px -20px; border-bottom: 1px solid #ddd; border-radius: 10px 10px 0 0; }
        .info-group { margin-bottom: 10px; }
        .label { font-weight: bold; color: #555; display: inline-block; width: 80px; }
        .photo-box { position: absolute; top: 60px; right: 20px; width: 100px; height: 130px; border: 1px solid #ddd; }
        .photo-box img { width: 100%; height: 100%; object-fit: cover; }
        .career-table { width: 100%; border-collapse: collapse; margin-top: 10px; font-size: 0.9rem; }
        .career-table th, .career-table td { border: 1px solid #ddd; padding: 5px; text-align: center; }
        .career-table th { background: #f1f3f5; }
        .checkbox-wrapper { margin-right: 10px; transform: scale(1.5); }
    </style>
    <script>
        function toggleAll(source) {
            checkboxes = document.getElementsByName('selected_ids');
            for(var i=0, n=checkboxes.length;i<n;i++) {
                checkboxes[i].checked = source.checked;
            }
        }
    </script>
    <div class="container">
        <div class="header">
            <h2 style="color:#003057; margin:0;">관리자 페이지</h2>
            <div>
                <a href="/download_excel" class="btn btn-excel">📊 엑셀 다운로드</a>
                <button onclick="document.getElementById('delete-form').submit();" class="btn btn-delete">🗑️ 선택 삭제</button>
                <a href="/logout" class="btn btn-logout">로그아웃</a>
            </div>
        </div>
        
        <form id="delete-form" action="/delete_selected" method="POST" onsubmit="return confirm('정말 삭제하시겠습니까?');">
            <div style="margin-bottom: 10px;">
                <input type="checkbox" onclick="toggleAll(this)"> 전체 선택
            </div>
    '''
    
    if not data:
        html += "<p style='text-align:center; padding:50px;'>접수된 지원서가 없습니다.</p>"
    
    for entry in data:
        photo_html = '<div class="photo-box" style="display:flex; align-items:center; justify-content:center; background:#eee; color:#aaa; font-size:0.8rem;">사진없음</div>'
        if entry['photo']:
            photo_html = f'<div class="photo-box"><a href="/view_photo/{entry["photo"]}" target="_blank"><img src="/view_photo/{entry["photo"]}"></a></div>'
            
        career_rows = ""
        for car in entry['career']:
            career_rows += f"<tr><td>{car['company']}</td><td>{car['start']}~{car['end']}</td><td>{car['role']}</td><td>{car['reason']}</td></tr>"
        if not career_rows:
            career_rows = "<tr><td colspan='4'>경력 사항 없음</td></tr>"

        html += f'''
        <div class="card">
            <div class="card-header">
                <div>
                    <input type="checkbox" name="selected_ids" value="{entry['id']}" class="checkbox-wrapper">
                    <span style="font-weight:bold; font-size:1.1rem;">{entry['info']['name']} ({entry['info']['birth']})</span>
                    <span style="color:#888; font-size:0.9rem; margin-left:10px;">접수: {entry['timestamp']}</span>
                </div>
            </div>
            {photo_html}
            <div style="width: 75%;">
                <div class="info-group"><span class="label">연락처</span> {entry['info']['phone']}</div>
                <div class="info-group"><span class="label">주소</span> {entry['info']['address']}</div>
                <div class="info-group"><span class="label">신체</span> {entry['body']['height']}cm / {entry['body']['weight']}kg / {entry['body']['vision']} / {entry['body']['shoes']} / {entry['body']['tshirt']}</div>
                <div class="info-group"><span class="label">근무조건</span> {entry['work_condition']['shift']} / {entry['work_condition']['posture']} / 잔업:{entry['work_condition']['overtime']} / 특근:{entry['work_condition']['holiday']}</div>
                <div class="info-group"><span class="label">희망일정</span> 입사희망: <span style="color:#0056b3; font-weight:bold;">{entry['work_condition']['start_date']}</span> (면접가능: {entry['work_condition']['interview_date']})</div>
                
                <table class="career-table">
                    <tr><th width="25%">회사명</th><th width="30%">기간</th><th width="20%">담당업무</th><th width="25%">퇴사사유</th></tr>
                    {career_rows}
                </table>
            </div>
        </div>
        '''
        
    html += '''
        </form>
    </div>
    '''
    return html

@app.route('/delete_selected', methods=['POST'])
def delete_selected():
    if not session.get('is_admin'):
        return redirect(url_for('login'))
        
    selected_ids = request.form.getlist('selected_ids')
    if not selected_ids:
        return "<script>alert('삭제할 항목을 선택해주세요.'); history.back();</script>"
        
    data = load_data()
    # 유지할 데이터만 필터링 (선택되지 않은 것들)
    new_data = [entry for entry in data if entry['id'] not in selected_ids]
    
    # 파일 삭제 (선택된 것들의 사진 파일)
    for entry in data:
        if entry['id'] in selected_ids and entry['photo']:
            try:
                os.remove(os.path.join(UPLOAD_DIR, entry['photo']))
            except:
                pass
                
    save_data(new_data)
    return redirect(url_for('master_view'))

@app.route('/download_excel')
def download_excel():
    if not session.get('is_admin'):
        return redirect(url_for('login'))
        
    data = load_data()
    wb = Workbook()
    ws = wb.active
    ws.title = "지원자 목록"
    
    # 스타일 정의
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    row_idx = 1
    
    for entry in data:
        # 1. 지원자별 헤더 (이름)
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
        cell = ws.cell(row=row_idx, column=1, value=f"[{entry['timestamp']}] {entry['info']['name']}")
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="003057", end_color="003057", fill_type="solid")
        cell.alignment = Alignment(vertical='center')
        row_idx += 1
        
        # 2. 사진 넣기 (A열 ~ B열 병합된 공간 확보)
        # 사진이 들어갈 공간 확보 (약 5줄 정도)
        photo_start_row = row_idx
        
        # 3. 데이터 입력 (Key - Value 형태)
        fields = [
            ("이름", entry['info']['name']),
            ("생년월일", entry['info']['birth']),
            ("연락처", entry['info']['phone']),
            ("이메일", entry['info']['email']),
            ("주소", entry['info']['address']),
            ("신체정보", f"{entry['body']['height']}cm / {entry['body']['weight']}kg"),
            ("상세사이즈", f"시력:{entry['body']['vision']} / 신발:{entry['body']['shoes']} / 티셔츠:{entry['body']['tshirt']}"),
            ("근무조건", f"{entry['work_condition']['shift']} / {entry['work_condition']['posture']}"),
            ("가능여부", f"잔업:{entry['work_condition']['overtime']} / 특근:{entry['work_condition']['holiday']}"),
            ("희망일정", f"면접:{entry['work_condition']['interview_date']} / 입사:{entry['work_condition']['start_date']}")
        ]
        
        # C열(항목명), D열(내용) 에 데이터 쓰기
        current_data_row = row_idx
        for key, value in fields:
            ws.cell(row=current_data_row, column=3, value=key).font = bold_font
            ws.cell(row=current_data_row, column=3, value=key).border = border_style
            ws.cell(row=current_data_row, column=3).alignment = center_align
            
            ws.cell(row=current_data_row, column=4, value=value).border = border_style
            ws.cell(row=current_data_row, column=4).alignment = Alignment(wrap_text=True, vertical='center')
            current_data_row += 1
            
        # 4. 경력사항 처리
        ws.cell(row=current_data_row, column=3, value="경력사항").font = bold_font
        ws.cell(row=current_data_row, column=3).border = border_style
        ws.cell(row=current_data_row, column=3).alignment = center_align
        
        career_text = ""
        if not entry['career']:
            career_text = "경력 없음"
        else:
            for c in entry['career']:
                career_text += f"[{c['company']}] {c['start']}~{c['end']} / {c['role']} / {c['reason']}\n"
        
        ws.cell(row=current_data_row, column=4, value=career_text.strip()).border = border_style
        ws.cell(row=current_data_row, column=4).alignment = Alignment(wrap_text=True, vertical='center')
        current_data_row += 1
        
        # 5. 사진 삽입 (A1:B10 영역 병합 후 삽입)
        # 병합
        ws.merge_cells(start_row=photo_start_row, start_column=1, end_row=current_data_row-1, end_column=2)
        img_cell = ws.cell(row=photo_start_row, column=1)
        img_cell.border = border_style
        
        if entry['photo']:
            try:
                img_path = os.path.join(UPLOAD_DIR, entry['photo'])
                if os.path.exists(img_path):
                    img = ExcelImage(img_path)
                    # 이미지 크기 조정 (약간 작게)
                    img.width = 150
                    img.height = 200
                    # 이미지 위치 (앵커)
                    ws.add_image(img, f"A{photo_start_row}")
                else:
                    img_cell.value = "이미지 파일 없음"
                    img_cell.alignment = center_align
            except Exception as e:
                img_cell.value = f"이미지 오류: {str(e)}"
        else:
            img_cell.value = "사진 없음"
            img_cell.alignment = center_align
            
        row_idx = current_data_row + 2 # 다음 지원자와 간격 띄우기

    # 컬럼 너비 조정
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 50

    excel_file = "applicants_export.xlsx"
    wb.save(excel_file)
    return send_file(excel_file, as_attachment=True)

@app.route('/view_photo/<filename>')
def view_photo(filename):
    return send_from_directory(UPLOAD_DIR, filename)

@app.route('/clear_data')
def clear_data():
    # 전체 삭제 (백업용) - 실제로는 잘 안쓰게 될 것임 (선택 삭제가 있어서)
    if not session.get('is_admin'):
        return redirect(url_for('login'))
    
    if os.path.exists(DATA_FILE):
        os.remove(DATA_FILE)
    if os.path.exists('data.xlsx'): # 구버전 파일 삭제
        os.remove('data.xlsx')
    if os.path.exists('data_html.txt'): # 구버전 파일 삭제
        os.remove('data_html.txt')
        
    # 업로드된 사진들도 삭제
    for f in os.listdir(UPLOAD_DIR):
        try:
            os.remove(os.path.join(UPLOAD_DIR, f))
        except: pass
        
    return redirect(url_for('master_view'))

if __name__ == '__main__':
    # SSL 인증서 경로 (서버 환경)
    cert_path = '/etc/letsencrypt/live/humetix.com/fullchain.pem'
    key_path = '/etc/letsencrypt/live/humetix.com/privkey.pem'

    if os.path.exists(cert_path) and os.path.exists(key_path):
        app.run(host='0.0.0.0', port=443, ssl_context=(cert_path, key_path))
    else:
        app.run(host='0.0.0.0', port=80, debug=True)