import os
import sys
from app import app
from models import db, Application
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, Alignment, Border, Side
import openpyxl

def generate_excel():
    with app.app_context():
        apps = Application.query.all()
        print(f"Found {len(apps)} applications.")
        data = [app.to_dict() for app in apps]
        
        wb = Workbook()
        ws = wb.active
        ws.title = "지원자 목록"
        
        # 스타일 정의
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        row_idx = 1
        UPLOAD_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')

        for entry in data:
            try:
                print(f"Processing {entry['info']['name']}...")
                # 1. 지원자별 헤더
                ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
                cell = ws.cell(row=row_idx, column=1, value=f"[{entry['timestamp']}] {entry['info']['name']}")
                cell.font = Font(bold=True, size=12, color="FFFFFF")
                cell.fill = openpyxl.styles.PatternFill(start_color="003057", end_color="003057", fill_type="solid")
                cell.alignment = Alignment(vertical='center')
                row_idx += 1
                
                # 2. 사진 공간 확보
                photo_start_row = row_idx
                
                # 3. 데이터 입력
                fields = [
                    ("이름", entry['info']['name']),
                    ("생년월일", entry['info']['birth']),
                    ("연락처", entry['info']['phone']),
                    ("이메일", entry['info']['email']),
                    ("주소", entry['info']['address']),
                    ("신체정보", f"{entry['body']['height']}cm / {entry['body']['weight']}kg"),
                    ("상세사이즈", f"시력:{entry['body']['vision']} / 신발:{entry['body']['shoes']} / 티셔츠:{entry['body']['tshirt']}"),
                    ("근무조건", f"{entry['work_condition']['shift']} / {entry['work_condition']['posture']}"),
                    ("가능여부", f"잔업:{entry['work_condition']['overtime']} / 특근:{entry['work_condition']['holiday']}"),
                    ("희망일정", f"면접:{entry['work_condition']['interview_date']} / 입사:{entry['work_condition']['start_date']}")
                ]
                
                current_data_row = row_idx
                for key, value in fields:
                    ws.cell(row=current_data_row, column=3, value=key).font = bold_font
                    ws.cell(row=current_data_row, column=3).border = border_style
                    ws.cell(row=current_data_row, column=3).alignment = center_align
                    ws.cell(row=current_data_row, column=4, value=value).border = border_style
                    ws.cell(row=current_data_row, column=4).alignment = Alignment(wrap_text=True, vertical='center')
                    current_data_row += 1
                    
                # 4. 경력사항
                ws.cell(row=current_data_row, column=3, value="경력사항").font = bold_font
                ws.cell(row=current_data_row, column=3).border = border_style
                ws.cell(row=current_data_row, column=3).alignment = center_align
                
                career_text = ""
                if not entry['career']:
                    career_text = "경력 없음"
                else:
                    for c in entry['career']:
                        if c:
                            career_text += f"[{c.get('company','')}] {c.get('start','')}~{c.get('end','')} / {c.get('role','')} / {c.get('reason','')}\n"
                
                ws.cell(row=current_data_row, column=4, value=career_text.strip()).border = border_style
                ws.cell(row=current_data_row, column=4).alignment = Alignment(wrap_text=True, vertical='center')
                current_data_row += 1
                
                # 5. 사진 삽입
                # Check merge range validity
                if photo_start_row > current_data_row - 1:
                     print(f"Merge range error: start={photo_start_row}, end={current_data_row-1}")

                ws.merge_cells(start_row=photo_start_row, start_column=1, end_row=current_data_row-1, end_column=2)
                img_cell = ws.cell(row=photo_start_row, column=1)
                img_cell.border = border_style
                
                if entry['photo']:
                    try:
                        img_path = os.path.join(UPLOAD_DIR, entry['photo'])
                        if os.path.exists(img_path):
                            img = ExcelImage(img_path)
                            img.width = 150
                            img.height = 200
                            ws.add_image(img, f"A{photo_start_row}")
                        else:
                            img_cell.value = "이미지 파일 없음"
                            img_cell.alignment = center_align
                    except Exception as e:
                        img_cell.value = f"이미지 오류: {str(e)}"
                        print(f"Image error: {e}")
                else:
                    img_cell.value = "사진 없음"
                    img_cell.alignment = center_align
                    
                row_idx = current_data_row + 2
                
            except Exception as e:
                print(f"Error processing entry {entry['id']}: {e}")
                import traceback
                traceback.print_exc()

        print("Excel generation successful!")

if __name__ == "__main__":
    generate_excel()
